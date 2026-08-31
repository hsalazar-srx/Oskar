"""
OSKAR — BOM browse + explosion endpoints (Slice A/B, ADR-012)

GET /api/v1/bom/{item_number}                 Single-level BOM browse (Slice A, B-1)
GET /api/v1/bom/{item_number}/indented         Multi-level explosion (Slice B, B-2)
GET /api/v1/bom/{item_number}/where-used       Where-used lookup (Slice B, B-3)
GET /api/v1/bom/{item_number}/export           TXT/CSV export (Slice F, I2-12)
POST /api/v1/bom/{item_number}/enrich          Supplier attributes (Slice F, I2-12)
"""

from __future__ import annotations

import io
from typing import Annotated, Any

import httpx
import openpyxl
from fastapi import APIRouter, Depends, Form, HTTPException, Query, Request, Response, UploadFile, status
from openpyxl import Workbook
from pydantic import BaseModel, Field
from sqlalchemy.ext.asyncio import AsyncSession

from src.adapters.erp.base import BOMNotFound
from src.adapters.erp.movex import MovexRestAdapter
from src.auth.dependencies import CurrentUser, get_current_user
from src.db import get_session
from src.routers.bulk_upload import BulkUploadSpec, parse_bulk_upload
from src.services.bom.browse import get_single_level_bom
from src.services.bom.compare import CompareOptions, diff_boms
from src.services.bom.comparisons import get_comparison, insert_comparison
from src.services.bom.customer_bom import CustomerLine, compare_customer_bom
from src.adapters.suppliers.chain import SupplierChain
from src.services.bom.enrich import enrich_bom_components
from src.services.bom.explode import assemble_where_used, build_bom_tree
from src.services.bom.export import UnsupportedExportFormat, export_bom
from src.services.bom.models import BOMCycleError, BOMHead, BOMTreeNode, WhereUsedLine
from src.services.bom.snapshots import get_snapshot

bom_router = APIRouter(prefix="/bom", tags=["bom"])


# ── Dependencies ─────────────────────────────────────────────────────────────

def _get_erp_adapter(request: Request) -> MovexRestAdapter:
    return request.app.state.erp_adapter


# ── Response models ───────────────────────────────────────────────────────────

class BOMLineResponse(BaseModel):
    sequence_number: int
    component_number: str
    description: str
    operation_number: int
    quantity: float
    unit_of_measure: str
    from_date: int
    to_date: int
    item_type: str | None
    status: str | None
    ref_des: list[str] | None
    customer_alias: str | None


class BOMHeadResponse(BaseModel):
    item_number: str
    structure_type: str
    facility: str
    description: str
    lines: list[BOMLineResponse]


def _to_response(head: BOMHead) -> BOMHeadResponse:
    return BOMHeadResponse(
        item_number=head.item_number,
        structure_type=head.structure_type,
        facility=head.facility,
        description=head.description,
        lines=[BOMLineResponse(**vars(line)) for line in head.lines],
    )


class BOMTreeNodeResponse(BaseModel):
    component_number: str
    description: str
    operation_number: int
    quantity: float
    cumulative_quantity: float
    item_type: str | None
    is_phantom: bool
    children: list["BOMTreeNodeResponse"] = []


BOMTreeNodeResponse.model_rebuild()


def _tree_to_response(node: BOMTreeNode) -> BOMTreeNodeResponse:
    return BOMTreeNodeResponse(
        component_number=node.component_number,
        description=node.description,
        operation_number=node.operation_number,
        quantity=node.quantity,
        cumulative_quantity=node.cumulative_quantity,
        item_type=node.item_type,
        is_phantom=node.is_phantom,
        children=[_tree_to_response(child) for child in node.children],
    )


class WhereUsedResponse(BaseModel):
    parent_item: str
    structure_type: str
    facility: str
    sequence_number: int
    component_number: str
    operation_number: int
    quantity: float
    unit_of_measure: str
    from_date: int
    to_date: int


def _where_used_to_response(line: WhereUsedLine) -> WhereUsedResponse:
    return WhereUsedResponse(**vars(line))


# ── Shared ERP error mapping ───────────────────────────────────────────────────

def _raise_for_erp_error(exc: Exception) -> None:
    """Map ERPAdapter exceptions to HTTP errors, matching src/routers/parts.py's
    established convention for MovexRestAdapter call sites."""
    if isinstance(exc, RuntimeError):
        if "circuit breaker" not in str(exc):
            raise exc
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="ERP system unavailable (circuit breaker open). Try again shortly.",
        )
    if isinstance(exc, httpx.HTTPStatusError):
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail=f"ERP returned unexpected status {exc.response.status_code}.",
        )
    if isinstance(exc, (httpx.ConnectError, httpx.TimeoutException)):
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail="ERP connection failed after retries.",
        )
    raise exc


# ── Slice D: BOM comparison engine ──────────────────────────────────────────────
# Declared BEFORE the "/{item_number}" routes below (Slice A/B) — "compare"
# and "comparisons" would otherwise be swallowed as an item_number path
# param, same reasoning as ecn_items.py's bulk-upload-before-{item_id} note.

class ComparisonResponse(BaseModel):
    id: str
    left_descriptor: dict[str, Any]
    right_descriptor: dict[str, Any]
    comparison_result: dict[str, Any]
    cost_impact: float | None
    risk_flags: list[str]
    created_by: str
    created_at: str


def _comparison_to_response(comp) -> ComparisonResponse:
    return ComparisonResponse(
        id=comp.id,
        left_descriptor=comp.left_descriptor,
        right_descriptor=comp.right_descriptor,
        comparison_result=comp.comparison_result,
        cost_impact=float(comp.cost_impact) if comp.cost_impact is not None else None,
        risk_flags=comp.risk_flags,
        created_by=comp.created_by,
        created_at=comp.created_at.isoformat(),
    )


@bom_router.get(
    "/comparisons/{comparison_id}",
    response_model=ComparisonResponse,
    summary="Fetch a saved BOM comparison (Slice D)",
)
async def get_bom_comparison(
    comparison_id: str,
    user: Annotated[CurrentUser, Depends(get_current_user)],
    session: Annotated[AsyncSession, Depends(get_session)],
) -> ComparisonResponse:
    comparison = await get_comparison(session, comparison_id)
    if comparison is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"No saved comparison found for id {comparison_id!r}.",
        )
    return _comparison_to_response(comparison)


_EXPORT_HEADER = ["Status", "Key", "Field", "Old Value", "New Value"]


def _comparison_result_to_workbook(comparison_result: dict[str, Any]) -> Workbook:
    """PLM parity: export always uses this FIXED field set, regardless of
    what CompareOptions.fields the comparison itself was restricted to on
    screen (ai/tasks/oskar-iteration-2.md Context: "Export to .xlsx only,
    always using the full fixed field set regardless of on-screen column
    visibility"). One row per (line, field_change) — a changed line with
    three field changes produces three export rows, one per changed field,
    so every value that differs is individually visible rather than
    collapsed into one cell."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM Comparison"
    ws.append(_EXPORT_HEADER)

    for line in comparison_result.get("added", []):
        ws.append(["Added", _line_display_key(line), "", "", _line_display_value(line)])
    for line in comparison_result.get("removed", []):
        ws.append(["Removed", _line_display_key(line), "", _line_display_value(line), ""])
    for changed in comparison_result.get("changed", []):
        key_display = "/".join(str(k) for k in changed.get("key", []))
        for fc in changed.get("field_changes", []):
            ws.append([
                "Changed", key_display, fc.get("field", ""),
                fc.get("old_value"), fc.get("new_value"),
            ])

    return wb


def _line_display_key(line: dict[str, Any]) -> str:
    for candidate in ("item_number", "component_number", "ipn"):
        if candidate in line:
            return str(line[candidate])
    return ""


def _line_display_value(line: dict[str, Any]) -> str:
    return ", ".join(f"{k}={v}" for k, v in line.items())


@bom_router.get(
    "/comparisons/{comparison_id}/export",
    summary="Export a saved BOM comparison to .xlsx (Slice D — fixed field set, PLM parity)",
)
async def export_bom_comparison(
    comparison_id: str,
    user: Annotated[CurrentUser, Depends(get_current_user)],
    session: Annotated[AsyncSession, Depends(get_session)],
):
    comparison = await get_comparison(session, comparison_id)
    if comparison is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"No saved comparison found for id {comparison_id!r}.",
        )

    wb = _comparison_result_to_workbook(comparison.comparison_result)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="bom-comparison-{comparison_id}.xlsx"'
        },
    )


class CompareSideDescriptor(BaseModel):
    """{type: erp|snapshot, ...identifying fields} — no "upload" here; upload
    sides go through POST /bom/compare/upload instead, which builds its
    customer-line list directly from the multipart file rather than via a
    descriptor round trip."""
    type: str
    item_number: str | None = None
    facility: str | None = None
    structure_type: str | None = "001"
    snapshot_id: str | None = None


class CompareOptionsBody(BaseModel):
    key: list[str] | None = None
    fields: list[str] | None = None


class CompareRequest(BaseModel):
    left: CompareSideDescriptor
    right: CompareSideDescriptor
    options: CompareOptionsBody = CompareOptionsBody()


async def _resolve_compare_side(
    descriptor: CompareSideDescriptor,
    erp: MovexRestAdapter,
    session: AsyncSession,
) -> list[dict[str, Any]]:
    """Descriptor -> plain line-dict list for diff_boms(). 'erp' calls
    get_single_level_bom (Slice A) and flattens BOMLine dataclasses to
    dicts; 'snapshot' fetches a bom_snapshots row (Slice D, migration 0026)
    and returns its stored lines as-is (already plain dicts, JSONB round
    trip)."""
    if descriptor.type == "erp":
        if not descriptor.item_number or not descriptor.facility:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="'erp' descriptor requires item_number and facility.",
            )
        head = await get_single_level_bom(
            erp, descriptor.item_number, descriptor.facility,
            structure_type=descriptor.structure_type or "001",
        )
        return [vars(line) for line in head.lines]

    if descriptor.type == "snapshot":
        if not descriptor.snapshot_id:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="'snapshot' descriptor requires snapshot_id.",
            )
        snapshot = await get_snapshot(session, descriptor.snapshot_id)
        if snapshot is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"No snapshot found for id {descriptor.snapshot_id!r}.",
            )
        return snapshot.lines

    raise HTTPException(
        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
        detail=f"Unsupported compare side type {descriptor.type!r}. Use 'erp' or 'snapshot'.",
    )


@bom_router.post(
    "/compare",
    response_model=ComparisonResponse,
    summary="Compare two BOMs (ERP item or saved snapshot) and save the result (Slice D)",
)
async def post_bom_compare(
    body: CompareRequest,
    user: Annotated[CurrentUser, Depends(get_current_user)],
    erp: Annotated[MovexRestAdapter, Depends(_get_erp_adapter)],
    session: Annotated[AsyncSession, Depends(get_session)],
) -> ComparisonResponse:
    """Resolve both descriptors to line lists, diff via compare.diff_boms,
    persist the result (bom_comparisons, migration 0027), and return it.

    Match key defaults to compare.py's own default
    (component_number, operation_number) when options.key is omitted —
    the ERP-vs-ERP default per D5. options.fields=None diffs every field
    common to both sides (compare.py's own default), i.e. the "no toggle
    applied" case.
    """
    try:
        left_lines = await _resolve_compare_side(body.left, erp, session)
        right_lines = await _resolve_compare_side(body.right, erp, session)
    except BOMNotFound as exc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(exc))
    except (RuntimeError, httpx.HTTPStatusError, httpx.ConnectError, httpx.TimeoutException) as exc:
        _raise_for_erp_error(exc)
        raise  # unreachable

    opts_kwargs: dict[str, Any] = {}
    if body.options.key is not None:
        opts_kwargs["key"] = tuple(body.options.key)
    if body.options.fields is not None:
        opts_kwargs["fields"] = tuple(body.options.fields)
    opts = CompareOptions(**opts_kwargs)

    diff = diff_boms(left_lines, right_lines, opts=opts)
    result_payload = {
        "added": diff.added,
        "removed": diff.removed,
        "changed": [
            {
                "key": list(c.key),
                "left": c.left,
                "right": c.right,
                "field_changes": [
                    {"field": fc.field, "old_value": fc.old_value, "new_value": fc.new_value}
                    for fc in c.field_changes
                ],
            }
            for c in diff.changed
        ],
        "unresolved": [
            {"side": u.side, "line": u.line, "reason": u.reason} for u in diff.unresolved
        ],
        "stats": vars(diff.stats),
    }

    saved = await insert_comparison(
        session,
        left_descriptor=body.left.model_dump(exclude_none=True),
        right_descriptor=body.right.model_dump(exclude_none=True),
        comparison_result=result_payload,
        created_by=user.username,
    )
    return _comparison_to_response(saved)


# ── Slice D: customer-BOM compare via file upload ───────────────────────────────
# Reuses BulkUploadSpec/parse_bulk_upload (src/routers/bulk_upload.py) per the
# plan's "reuse ecn_items.py bulk constants" note. Canonical customer-BOM
# column shape (parity with PLM's upload flow, ai/tasks/oskar-iteration-2.md
# Context): IPN, CPN, MFR1/MPN1, MFR2/MPN2 (repeated manufacturer/part-number
# pairs collapse into CustomerLine.mpn[]/mfr[] arrays), Designator,
# Description, Quantity. Footprint is a mappable column but dropped before
# comparison (dead in PLM too, per the plan) — simply absent from column_map.

_CUSTOMER_BOM_UPLOAD_SPEC = BulkUploadSpec(
    template_name="customer BOM compare template",
    required_columns={"IPN", "MFR1", "MPN1", "Quantity"},
    column_map={
        "ipn": "ipn",
        "cpn": "cpn",
        "mfr1": "mfr_1",
        "mpn1": "mpn_1",
        "mfr2": "mfr_2",
        "mpn2": "mpn_2",
        "designator": "designator",
        "description": "description",
        "quantity": "quantity",
    },
    row_key_field="ipn",
)


async def _strip_leading_title_row_xlsx(file: UploadFile, spec: BulkUploadSpec) -> UploadFile:
    """Work around parse_bulk_upload/_parse_xlsx's fixed "first non-blank
    row is the header" rule (src/routers/bulk_upload.py — outside this
    slice's file boundary, reused as-is, not modified) for the one real
    case Slice 0's customer_bom.xlsx fixture was deliberately built to
    exercise: a title row ("Customer BOM Export — generated ...") above the
    real header row (its own docstring: "PLM's upload flow lets the user
    pick which row is the header... this fixture exercises that row is not
    always row 1"). Rather than silently failing that fixture or extending
    the shared helper, this endpoint-local pre-pass finds the first row
    that actually contains every one of spec.required_columns and rewrites
    the workbook to start there before handing bytes to parse_bulk_upload.
    A no-op (returns the original UploadFile untouched) for any file whose
    first row already IS the header, or for non-xlsx content types."""
    content_type = (file.content_type or "").lower().split(";")[0].strip()
    if content_type not in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ):
        return file

    raw = await file.read()
    await file.seek(0)

    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    required_normalised = {c.strip().lower() for c in spec.required_columns}

    rows = list(ws.iter_rows(values_only=True))
    header_row_idx = None
    for idx, row in enumerate(rows):
        cells = {str(c).strip().lower() for c in row if c is not None}
        if required_normalised <= cells:
            header_row_idx = idx
            break
    wb.close()

    if header_row_idx is None or header_row_idx == 0:
        return file  # no title row to strip (or headers never found — let
        # parse_bulk_upload's own missing-columns 422 report that clearly)

    new_wb = Workbook()
    new_ws = new_wb.active
    for row in rows[header_row_idx:]:
        new_ws.append(row)
    buf = io.BytesIO()
    new_wb.save(buf)
    buf.seek(0)
    return UploadFile(filename=file.filename, file=buf, headers=file.headers)


class _CustomerBomUploadRow(BaseModel):
    """Per-row validation (plan: "422 with row numbers on bad rows"). Only
    quantity presence is enforced here — quantity does NOT have to be
    numeric (defect (b) fix, compare.py module docstring: a non-numeric
    quantity like "N/A" is valid data, not an error, and diffs correctly
    via string-equality fallback). ipn is guaranteed non-blank already by
    parse_bulk_upload's row_key_field mechanism, so is not re-validated
    here."""
    ipn: str = Field(..., min_length=1)
    cpn: str | None = None
    mfr_1: str | None = None
    mpn_1: str | None = None
    mfr_2: str | None = None
    mpn_2: str | None = None
    designator: str | None = None
    description: str | None = None
    quantity: str = Field(..., min_length=1)


def _row_to_customer_line(row: _CustomerBomUploadRow) -> CustomerLine:
    mpn = [m for m in (row.mpn_1, row.mpn_2) if m]
    mfr = [m for m in (row.mfr_1, row.mfr_2) if m]
    return CustomerLine(
        cpn=row.cpn,
        mpn=mpn,
        mfr=mfr,
        designator=row.designator,
        description=row.description,
        quantity=row.quantity,
        ipn=row.ipn,
    )


@bom_router.post(
    "/compare/upload",
    response_model=ComparisonResponse,
    summary="Compare an uploaded customer BOM (xlsx/csv) against an ERP item (Slice D, I2-2)",
)
async def post_bom_compare_upload(
    user: Annotated[CurrentUser, Depends(get_current_user)],
    erp: Annotated[MovexRestAdapter, Depends(_get_erp_adapter)],
    session: Annotated[AsyncSession, Depends(get_session)],
    file: UploadFile,
    item_number: Annotated[str, Form(description="Oskar/Movex item to compare the upload against")],
    facility: Annotated[str, Form(description="Movex facility (MPDHED.FACI)")] = "D",
) -> ComparisonResponse:
    """Parse the uploaded customer BOM, resolve each line to an Oskar
    item_number via src/services/bom/customer_bom.py's resolution rule (CPN
    alias through ERPAdapter.lookup_by_alias first, then MPN through
    item_mpns — see compare_customer_bom's docstring for why CPN takes
    priority), diff the resolved set against the ERP item's current
    single-level BOM, persist, and return the result. Lines resolving via
    neither path land in the response's `unresolved` bucket rather than
    silently vanishing from the comparison.

    422 (with per-row detail from parse_bulk_upload / row validation) on
    unparseable files, missing required columns, or an empty data set —
    same guard sequence as ecn_items.py's bulk endpoints.
    """
    file = await _strip_leading_title_row_xlsx(file, _CUSTOMER_BOM_UPLOAD_SPEC)
    raw_rows = await parse_bulk_upload(file, _CUSTOMER_BOM_UPLOAD_SPEC)

    validated_rows: list[_CustomerBomUploadRow] = []
    row_errors: list[str] = []
    for idx, raw in enumerate(raw_rows, start=1):
        try:
            validated_rows.append(_CustomerBomUploadRow(**raw))
        except Exception as exc:
            row_errors.append(f"Row {idx} ({raw.get('ipn', '?')}): {exc}")
    if row_errors:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="; ".join(row_errors),
        )

    customer_lines = [_row_to_customer_line(row) for row in validated_rows]

    try:
        erp_head = await get_single_level_bom(erp, item_number, facility)
    except BOMNotFound as exc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(exc))
    except (RuntimeError, httpx.HTTPStatusError, httpx.ConnectError, httpx.TimeoutException) as exc:
        _raise_for_erp_error(exc)
        raise  # unreachable

    erp_lines = [
        {"item_number": line.component_number, **vars(line)} for line in erp_head.lines
    ]

    # customer_bom.py's resolve_cpn/resolve_mpn contract is SYNCHRONOUS by
    # design (pure-logic module — see its docstring: DB/HTTP wiring happens
    # at the router layer). lookup_by_alias is a real async HTTP call, so
    # every CPN is resolved up front here (one lookup_by_alias call per
    # distinct CPN, not per line) into a plain dict, then handed to
    # compare_customer_bom as a synchronous dict.get closure.
    distinct_cpns = {ln.cpn for ln in customer_lines if ln.cpn}
    cpn_to_item: dict[str, str] = {}
    for cpn in distinct_cpns:
        # MITPOP reverse-alias lookup (same mechanism as GET /parts/alias).
        # Only a FULL MATCH (exactly one candidate) resolves — an ambiguous
        # (partial_match) or absent (no_match) alias is not a safe automatic
        # resolution and instead falls through to MPN, then to unresolved.
        candidates = await erp.lookup_by_alias(cpn)
        if len(candidates) == 1:
            cpn_to_item[cpn] = str(candidates[0]["ITNO"]).strip()

    # MPN resolution for this endpoint's specific compare-against-one-item
    # shape: a customer MPN resolves only if it matches an MPN already
    # present in the target item's OWN current BOM lines (no general
    # cross-catalogue item_mpns query is issued here — that would let a
    # customer file resolve to items entirely unrelated to the compare
    # target, which is not what "compare this upload against item X"
    # means). Exact-match on component_number text, since Slice A's browse
    # layer does not expose manufacturer_canonical per line for a synonym-
    # aware match.
    mpn_to_item: dict[str, str] = {
        str(ln.get("component_number", "")).strip().upper(): ln["item_number"]
        for ln in erp_lines
    }

    def _resolve_cpn(cpn: str) -> str | None:
        return cpn_to_item.get(cpn)

    def _resolve_mpn(mpn: str) -> str | None:
        return mpn_to_item.get(mpn.strip().upper())

    result = compare_customer_bom(
        customer_lines,
        erp_lines,
        resolve_cpn=_resolve_cpn,
        resolve_mpn=_resolve_mpn,
        opts=CompareOptions(key=("item_number",), fields=("quantity",)),
    )

    result_payload = {
        "added": result.added,
        "removed": result.removed,
        "changed": [
            {
                "key": list(c.key),
                "left": c.left,
                "right": c.right,
                "field_changes": [
                    {"field": fc.field, "old_value": fc.old_value, "new_value": fc.new_value}
                    for fc in c.field_changes
                ],
            }
            for c in result.changed
        ],
        "unresolved": [
            {"side": u.side, "line": u.line, "reason": u.reason} for u in result.unresolved
        ],
        "stats": vars(result.stats),
    }

    saved = await insert_comparison(
        session,
        left_descriptor={"type": "upload", "filename": file.filename},
        right_descriptor={"type": "erp", "item_number": item_number, "facility": facility},
        comparison_result=result_payload,
        created_by=user.username,
    )
    return _comparison_to_response(saved)


# ── Slice A: single-level browse ──────────────────────────────────────────────

@bom_router.get(
    "/{item_number}",
    response_model=BOMHeadResponse,
    summary="Single-level BOM browse (Slice A, B-1)",
)
async def get_bom(
    item_number: str,
    user: Annotated[CurrentUser, Depends(get_current_user)],
    erp: Annotated[MovexRestAdapter, Depends(_get_erp_adapter)],
    facility: Annotated[str, Query(max_length=5, description="Movex facility (MPDHED.FACI)")] = "D",
    structure_type: Annotated[str, Query(max_length=3, description="Movex structure type (MPDHED.STRT)")] = "001",
    bom_type: Annotated[str, Query(max_length=1, description="'M' = manufacturing BOM (default)")] = "M",
    effective_on: Annotated[str | None, Query(description="YYYYMMDD — optional as-of date passed to the ERP call")] = None,
    include_expired: Annotated[bool, Query(description="Include lines whose to_date has already passed")] = False,
) -> BOMHeadResponse:
    """Fetch a single-level BOM (MPDHED head + MPDMAT lines) for an item.

    Effectivity-filtered (to_date >= today, or >= effective_on when provided)
    unless include_expired=true. Lines are returned in MSEQ order.
    """
    try:
        head = await get_single_level_bom(
            erp,
            item_number,
            facility,
            structure_type=structure_type,
            bom_type=bom_type,
            effective_on=effective_on,
            include_expired=include_expired,
        )
    except BOMNotFound:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"No BOM found for item {item_number!r}.",
        )
    except (RuntimeError, httpx.HTTPStatusError, httpx.ConnectError, httpx.TimeoutException) as exc:
        _raise_for_erp_error(exc)
        raise  # unreachable — _raise_for_erp_error always raises

    return _to_response(head)


# ── Slice B: multi-level explosion ────────────────────────────────────────────

@bom_router.get(
    "/{item_number}/indented",
    response_model=BOMTreeNodeResponse,
    summary="Multi-level (indented) BOM explosion (Slice B, B-2)",
)
async def get_bom_indented(
    item_number: str,
    user: Annotated[CurrentUser, Depends(get_current_user)],
    erp: Annotated[MovexRestAdapter, Depends(_get_erp_adapter)],
    facility: Annotated[str, Query(max_length=5, description="Movex facility (MPDHED.FACI)")] = "D",
    structure_type: Annotated[str, Query(max_length=3, description="Movex structure type (MPDHED.STRT)")] = "001",
    depth: Annotated[int, Query(ge=1, le=12, description="Maximum explosion depth")] = 12,
) -> BOMTreeNodeResponse:
    """Fetch the multi-level explosion tree for an item.

    B-2 returns a flat, depth-first record list (recursive CTE over MPDMAT —
    PDZ100MI is broken in M3 and is not an option); this endpoint assembles it
    into a tree client-side via src/services/bom/explode.py.

    Note: B-2's own recursive-CTE performance against a large real multi-level
    UAT item (<2s target, ADR-012 Decision 4) is an external checkpoint owned
    by the movex-rest-api team — it cannot be validated from Oskar's side and
    is not exercised by this endpoint's tests.
    """
    try:
        payload = await erp.get_bom_indented(
            item_number,
            facility,
            structure_type=structure_type,
            max_depth=depth,
        )
    except (RuntimeError, httpx.HTTPStatusError, httpx.ConnectError, httpx.TimeoutException) as exc:
        _raise_for_erp_error(exc)
        raise  # unreachable

    try:
        data = payload.get("data", payload)
        tree = build_bom_tree(item_number, data.get("records", []), max_depth=depth)
    except BOMCycleError as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Could not assemble BOM tree for {item_number!r}: {exc}",
        )

    return _tree_to_response(tree)


# ── Slice B: where-used ────────────────────────────────────────────────────────

@bom_router.get(
    "/{item_number}/where-used",
    response_model=list[WhereUsedResponse],
    summary="Where-used lookup: assemblies that consume this component (Slice B, B-3)",
)
async def get_where_used(
    item_number: str,
    user: Annotated[CurrentUser, Depends(get_current_user)],
    erp: Annotated[MovexRestAdapter, Depends(_get_erp_adapter)],
    facility: Annotated[str, Query(max_length=5, description="Movex facility (MPDMAT.FACI)")] = "D",
    effective_on: Annotated[str | None, Query(description="YYYYMMDD — optional as-of date passed to the ERP call")] = None,
) -> list[WhereUsedResponse]:
    """List every parent assembly that consumes item_number as a component.

    An empty result is a legitimate "used nowhere" answer, not a 404.
    """
    try:
        payload = await erp.get_where_used(item_number, facility, effective_on=effective_on)
    except (RuntimeError, httpx.HTTPStatusError, httpx.ConnectError, httpx.TimeoutException) as exc:
        _raise_for_erp_error(exc)
        raise  # unreachable

    lines = assemble_where_used(payload)
    return [_where_used_to_response(line) for line in lines]


# ── Slice F: TXT/CSV export (I2-12) ──────────────────────────────────────────

@bom_router.get(
    "/{item_number}/export",
    summary="Export a single-level BOM as TXT or CSV (Slice F, I2-12)",
    response_class=Response,
)
async def export_bom_endpoint(
    item_number: str,
    user: Annotated[CurrentUser, Depends(get_current_user)],
    erp: Annotated[MovexRestAdapter, Depends(_get_erp_adapter)],
    format: Annotated[str, Query(description="'csv' (default) or 'txt'")] = "csv",
    facility: Annotated[str, Query(max_length=5, description="Movex facility (MPDHED.FACI)")] = "D",
    structure_type: Annotated[str, Query(max_length=3, description="Movex structure type (MPDHED.STRT)")] = "001",
    bom_type: Annotated[str, Query(max_length=1, description="'M' = manufacturing BOM (default)")] = "M",
    effective_on: Annotated[str | None, Query(description="YYYYMMDD — optional as-of date passed to the ERP call")] = None,
    include_expired: Annotated[bool, Query(description="Include lines whose to_date has already passed")] = False,
) -> Response:
    """Export the same single-level BOM that GET /bom/{item_number} returns.

    Deliberately reuses get_single_level_bom rather than re-fetching
    differently: an export that disagreed with what the browser shows for the
    same item would be worse than no export. The effectivity/expiry query
    params mirror the browse endpoint one-for-one for the same reason.

    xlsx is not offered here — comparison xlsx export lives at
    /bom/comparisons/{id}/export (Slice D). Asking for it returns 422 rather
    than silently handing back a CSV.
    """
    try:
        head = await get_single_level_bom(
            erp,
            item_number,
            facility,
            structure_type=structure_type,
            bom_type=bom_type,
            effective_on=effective_on,
            include_expired=include_expired,
        )
    except BOMNotFound:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"No BOM found for item {item_number!r}.",
        )
    except (RuntimeError, httpx.HTTPStatusError, httpx.ConnectError, httpx.TimeoutException) as exc:
        _raise_for_erp_error(exc)
        raise  # unreachable — _raise_for_erp_error always raises

    try:
        content, media_type, ext = export_bom(head, format)
    except UnsupportedExportFormat as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=str(exc),
        )

    return Response(
        content=content,
        media_type=media_type,
        headers={
            "Content-Disposition": f'attachment; filename="bom-{item_number}.{ext}"'
        },
    )


# ── Slice F: supplier attribute enrichment (I2-12) ───────────────────────────

# Hard ceiling on the per-request lookup cap. The cap protects a SHARED daily
# API budget (element14 and DigiKey are 1,000/day each), so a caller must not
# be able to raise it arbitrarily from a query param — that would let one
# request drain the budget for everyone.
_MAX_LIVE_LOOKUP_CAP = 200

# Statuses that mean "this result is not the whole picture, run it again"
# — as opposed to no_mpn/not_found, which are real findings about the data.
_INCOMPLETE_STATUSES = {"cap_reached", "lookup_failed"}


class EnrichedComponentResponse(BaseModel):
    sequence_number: int
    component_number: str
    description: str
    mpn: str | None
    status: str
    attributes: dict[str, Any] = {}


class BOMEnrichResponse(BaseModel):
    """Enrichment result for one BOM.

    `summary` counts components by status and `incomplete` says outright
    whether a re-run would add anything — so a client does not need to know
    which statuses imply "budget spent" or "supplier unreachable".
    """

    item_number: str
    components: list[EnrichedComponentResponse]
    summary: dict[str, int]
    incomplete: bool


@bom_router.post(
    "/{item_number}/enrich",
    response_model=BOMEnrichResponse,
    summary="Enrich BOM components with supplier attributes (Slice F, I2-12)",
)
async def enrich_bom(
    item_number: str,
    user: Annotated[CurrentUser, Depends(get_current_user)],
    erp: Annotated[MovexRestAdapter, Depends(_get_erp_adapter)],
    session: Annotated[AsyncSession, Depends(get_session)],
    request: Request,
    facility: Annotated[str, Query(max_length=5, description="Movex facility (MPDHED.FACI)")] = "D",
    structure_type: Annotated[str, Query(max_length=3, description="Movex structure type (MPDHED.STRT)")] = "001",
    live_lookup_cap: Annotated[
        int | None,
        Query(
            ge=1,
            le=_MAX_LIVE_LOOKUP_CAP,
            description=(
                "Max distinct MPNs to look up in this request. Omit for the "
                "service default. Capped to protect a shared daily API budget."
            ),
        ),
    ] = None,
) -> BOMEnrichResponse:
    """Look up supplier attributes for every component on a BOM.

    Cache-first and capped — see src/services/bom/enrich.py for why the cap
    is the central design constraint rather than a nicety. Components that
    could not be enriched are returned with a status explaining why, never
    dropped, so the gaps stay actionable.

    POST rather than GET: this can spend real, limited API quota, which is a
    side effect a GET should not have (and which caches/prefetchers would
    happily trigger on their own).
    """
    try:
        head = await get_single_level_bom(
            erp, item_number, facility, structure_type=structure_type
        )
    except BOMNotFound:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"No BOM found for item {item_number!r}.",
        )
    except (RuntimeError, httpx.HTTPStatusError, httpx.ConnectError, httpx.TimeoutException) as exc:
        _raise_for_erp_error(exc)
        raise  # unreachable

    chain = SupplierChain(session, getattr(request.app.state, "supplier_adapters", []))
    components = await enrich_bom_components(
        session, head, chain, live_lookup_cap=live_lookup_cap
    )

    summary: dict[str, int] = {}
    for component in components:
        summary[component.status] = summary.get(component.status, 0) + 1

    return BOMEnrichResponse(
        item_number=item_number,
        components=[
            EnrichedComponentResponse(
                sequence_number=c.sequence_number,
                component_number=c.component_number,
                description=c.description,
                mpn=c.mpn,
                status=c.status,
                attributes=c.attributes,
            )
            for c in components
        ],
        summary=summary,
        incomplete=any(s in _INCOMPLETE_STATUSES for s in summary),
    )
