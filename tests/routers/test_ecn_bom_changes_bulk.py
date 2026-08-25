"""
OSKAR — Bulk BOM-change upload endpoint tests (Slice E, I2-6)

POST /api/v1/ecn/{ecn_id}/bom-changes/bulk (multipart/form-data)

Stargile parity: UploadECNBoMs. ECN-wide, multi-item — verified against the
real Stargile source (2026-08-11,
c:/Projects/SuperTool/Stargile_Source_Code/.../ecn/upload/rules/UploadECNBoMs.java):
each uploaded row carries its own zecnln (ECN line number) / prno (parent
item), i.e. an upload can spread across many items on the same ECN, the same
shape as routing's bulk upload (ecn_routing.py's /routing/bulk) — NOT scoped
to a single pre-selected item. This corrects an earlier draft of this
endpoint that read Stargile's upload as per-item; verified wrong against the
actual UploadECNBoMs.java column layout before landing.

Follows the exact S9-8 pattern (BulkUploadSpec/parse_bulk_upload -> in-file
dup-check -> Pydantic row validation -> service call) established by
ecn_routing.py's bulk endpoint — see tests/routers/test_ecn_routing_bulk.py,
which this file mirrors structurally.
"""
from __future__ import annotations

import io
from datetime import datetime, timezone
from unittest.mock import AsyncMock, patch

import openpyxl
import pytest
from fastapi.testclient import TestClient

from src.auth.dependencies import CurrentUser, get_current_user
from src.db import get_session
from src.main import app
from src.routers.ecn_bom import _get_erp_adapter
from src.services.ecn import ECNService
from src.services.ecn.models import BOMChangeResponse, ECNNotFound, ECNValidationError

_NOW = datetime(2026, 8, 11, 10, 0, 0, tzinfo=timezone.utc)

_ENGINEER = CurrentUser(
    username="eng_user",
    display_name="Engineer",
    email="eng@scanfil.com",
    groups=["ecn-initiator"],
    jti="test-jti-bom-bulk-001",
)

_ECN_ID = "ecn-uuid-bom-bulk-001"

_HEADERS = [
    "Item No", "Component Number", "Change Type", "Quantity", "Unit of Measure",
    "Operation Number", "From Date", "Old From Date", "Old Quantity",
]


def _make_change(
    item_id: str, component_number: str, change_type: str = "ADD", quantity: float = 4.0,
    item_number: str | None = None,
) -> BOMChangeResponse:
    return BOMChangeResponse(
        id=f"bc-{item_id}-{component_number}",
        ecn_id=_ECN_ID,
        parent_item_number=item_number or "LF100001",
        change_type=change_type,
        component_number=component_number,
        quantity=quantity,
        unit_of_measure="EA",
        operation_number=10,
        sequence_number=None,
        from_date=20260901,
        to_date=None,
        bom_type="M",
        notes=None,
        old_quantity=None,
        old_operation_number=None,
        old_from_date=None,
        old_to_date=None,
        circuit_refs_old=None,
        circuit_refs_new=None,
        snapshot_id=None,
        movex_snapshot_at_review=None,
        created_at=_NOW,
        ecn_item_id=item_id,
        item_number=item_number,
    )


def _make_xlsx(rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(_HEADERS)
    for row in rows:
        ws.append([row.get(h, "") for h in _HEADERS])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_csv(rows: list[dict]) -> bytes:
    lines = [",".join(_HEADERS)]
    for row in rows:
        lines.append(",".join(str(row.get(h, "")) for h in _HEADERS))
    return "\n".join(lines).encode("utf-8")


class _NoItemsSession:
    """Minimal session stub for the ADR-014 on-ECN-parent lookup.

    The bulk route now runs one `SELECT item_number FROM ecn_items` to work
    out which parents are already on the ECN (those skip the Movex check).
    ECNService is mocked per test, so this is the only real query the route
    makes. Returning no rows means every parent is treated as off-ECN, which
    is what exercises the new Movex-existence path.
    """

    async def execute(self, *_args, **_kwargs):
        return []


class _StubERPAdapter:
    """Every parent resolves — the happy path. Tests that need a missing
    parent patch this per-test."""

    async def get_item(self, item_number: str) -> dict:
        return {"itno": item_number}


@pytest.fixture()
def client():
    app.dependency_overrides[get_current_user] = lambda: _ENGINEER
    app.dependency_overrides[get_session] = lambda: _NoItemsSession()
    app.dependency_overrides[_get_erp_adapter] = lambda: _StubERPAdapter()
    yield TestClient(app)
    app.dependency_overrides.clear()


# ── Happy path ────────────────────────────────────────────────────────────────

class TestBulkBomChangeHappyPath:
    def test_multi_item_upload_returns_201(self, client):
        """Rows spread across two different items in one file — the same
        multi-item shape as routing's bulk upload, matching Stargile's
        UploadECNBoMs (verified: prno is a per-row field, not a URL param)."""
        rows = [
            {"Item No": "LFAM050001", "Component Number": "LF200010", "Change Type": "ADD",
             "Quantity": "4.0", "Unit of Measure": "EA", "Operation Number": "10", "From Date": "20260901"},
            {"Item No": "LFAM050002", "Component Number": "LF200020", "Change Type": "ADD",
             "Quantity": "2.0", "Unit of Measure": "EA", "Operation Number": "20", "From Date": "20260901"},
        ]
        xlsx_bytes = _make_xlsx(rows)
        created = [
            _make_change("item-LFAM050001", "LF200010", item_number="LFAM050001"),
            _make_change("item-LFAM050002", "LF200020", item_number="LFAM050002"),
        ]

        with patch.object(ECNService, "bulk_create_bom_changes", new=AsyncMock(return_value=created)):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/bom-changes/bulk",
                files={"file": ("bom.xlsx", xlsx_bytes,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 201
        data = resp.json()
        assert len(data) == 2
        assert data[0]["item_number"] == "LFAM050001"
        assert data[1]["item_number"] == "LFAM050002"

    def test_single_item_multi_row_upload_returns_201(self, client):
        """Many BOM-change rows for the SAME item — the common case."""
        rows = [
            {"Item No": "LFAM050001", "Component Number": "LF200010", "Change Type": "ADD",
             "Quantity": "4.0", "Operation Number": "10", "From Date": "20260901"},
            {"Item No": "LFAM050001", "Component Number": "LF200020", "Change Type": "ADD",
             "Quantity": "2.0", "Operation Number": "20", "From Date": "20260901"},
        ]
        xlsx_bytes = _make_xlsx(rows)
        created = [
            _make_change("item-LFAM050001", "LF200010", item_number="LFAM050001"),
            _make_change("item-LFAM050001", "LF200020", item_number="LFAM050001"),
        ]

        with patch.object(ECNService, "bulk_create_bom_changes", new=AsyncMock(return_value=created)):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/bom-changes/bulk",
                files={"file": ("bom.xlsx", xlsx_bytes,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 201
        assert len(resp.json()) == 2

    def test_csv_upload_returns_201(self, client):
        rows = [{"Item No": "LFAM050001", "Component Number": "LF200010", "Change Type": "ADD",
                 "Quantity": "4.0", "Operation Number": "10", "From Date": "20260901"}]
        csv_bytes = _make_csv(rows)
        created = [_make_change("item-LFAM050001", "LF200010", item_number="LFAM050001")]

        with patch.object(ECNService, "bulk_create_bom_changes", new=AsyncMock(return_value=created)):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/bom-changes/bulk",
                files={"file": ("bom.csv", csv_bytes, "text/csv")},
            )

        assert resp.status_code == 201
        assert len(resp.json()) == 1

    def test_change_type_row_with_old_from_date_returns_201(self, client):
        rows = [{"Item No": "LFAM050001", "Component Number": "LF200010", "Change Type": "CHANGE",
                 "Quantity": "6.0", "Operation Number": "10", "From Date": "20260901",
                 "Old From Date": "20240101", "Old Quantity": "4.0"}]
        xlsx_bytes = _make_xlsx(rows)
        created = [_make_change("item-LFAM050001", "LF200010", change_type="CHANGE", quantity=6.0,
                                 item_number="LFAM050001")]

        with patch.object(ECNService, "bulk_create_bom_changes", new=AsyncMock(return_value=created)):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/bom-changes/bulk",
                files={"file": ("bom.xlsx", xlsx_bytes,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 201


# ── Validation errors ─────────────────────────────────────────────────────────

class TestBulkBomChangeValidationErrors:
    def test_wrong_content_type_returns_422(self, client):
        resp = client.post(
            f"/api/v1/ecn/{_ECN_ID}/bom-changes/bulk",
            files={"file": ("bom.txt", b"some text", "text/plain")},
        )
        assert resp.status_code == 422

    def test_missing_required_columns_returns_422(self, client):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Wrong Col A", "Wrong Col B"])
        ws.append(["val1", "val2"])
        buf = io.BytesIO()
        wb.save(buf)

        with patch.object(ECNService, "bulk_create_bom_changes", new=AsyncMock()):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/bom-changes/bulk",
                files={"file": ("bad.xlsx", buf.getvalue(),
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 422

    def test_empty_file_returns_422(self, client):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(_HEADERS)
        buf = io.BytesIO()
        wb.save(buf)

        with patch.object(ECNService, "bulk_create_bom_changes", new=AsyncMock()):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/bom-changes/bulk",
                files={"file": ("empty.xlsx", buf.getvalue(),
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 422

    def test_invalid_change_type_returns_422(self, client):
        rows = [{"Item No": "LFAM050001", "Component Number": "LF200010", "Change Type": "BOGUS",
                 "Quantity": "4.0", "Operation Number": "10", "From Date": "20260901"}]
        xlsx_bytes = _make_xlsx(rows)

        with patch.object(ECNService, "bulk_create_bom_changes", new=AsyncMock()):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/bom-changes/bulk",
                files={"file": ("bom.xlsx", xlsx_bytes,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 422

    def test_change_type_without_old_from_date_returns_422(self, client):
        rows = [{"Item No": "LFAM050001", "Component Number": "LF200010", "Change Type": "CHANGE",
                 "Quantity": "6.0", "Operation Number": "10", "From Date": "20260901"}]  # no Old From Date
        xlsx_bytes = _make_xlsx(rows)

        with patch.object(
            ECNService, "bulk_create_bom_changes",
            new=AsyncMock(side_effect=ECNValidationError(
                "Row 1: old_from_date is required for change_type CHANGE/DELETE"
            )),
        ):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/bom-changes/bulk",
                files={"file": ("bom.xlsx", xlsx_bytes,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 422

    def test_duplicate_item_component_operation_triple_returns_409(self, client):
        rows = [
            {"Item No": "LFAM050001", "Component Number": "LF200010", "Change Type": "ADD",
             "Quantity": "4.0", "Operation Number": "10", "From Date": "20260901"},
            {"Item No": "LFAM050001", "Component Number": "LF200010", "Change Type": "ADD",
             "Quantity": "8.0", "Operation Number": "10", "From Date": "20260901"},
        ]
        xlsx_bytes = _make_xlsx(rows)

        resp = client.post(
            f"/api/v1/ecn/{_ECN_ID}/bom-changes/bulk",
            files={"file": ("bom.xlsx", xlsx_bytes,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

        assert resp.status_code == 409

    def test_same_component_different_items_not_a_duplicate(self, client):
        """The same (component_number, operation_number) pair on TWO
        different items is not a duplicate — the batch key includes
        item_number, matching routing's own (item_number, operation_number)
        dup-check convention."""
        rows = [
            {"Item No": "LFAM050001", "Component Number": "LF200010", "Change Type": "ADD",
             "Quantity": "4.0", "Operation Number": "10", "From Date": "20260901"},
            {"Item No": "LFAM050002", "Component Number": "LF200010", "Change Type": "ADD",
             "Quantity": "4.0", "Operation Number": "10", "From Date": "20260901"},
        ]
        xlsx_bytes = _make_xlsx(rows)
        created = [
            _make_change("item-LFAM050001", "LF200010", item_number="LFAM050001"),
            _make_change("item-LFAM050002", "LF200010", item_number="LFAM050002"),
        ]

        with patch.object(ECNService, "bulk_create_bom_changes", new=AsyncMock(return_value=created)):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/bom-changes/bulk",
                files={"file": ("bom.xlsx", xlsx_bytes,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 201
        assert len(resp.json()) == 2


# ── ECN / item state guards ──────────────────────────────────────────────────

class TestBulkBomChangeStateGuards:
    def _one_row_xlsx(self) -> bytes:
        return _make_xlsx([{"Item No": "LFAM050001", "Component Number": "LF200010",
                             "Change Type": "ADD", "Quantity": "4.0",
                             "Operation Number": "10", "From Date": "20260901"}])

    def test_ecn_not_found_returns_404(self, client):
        with patch.object(
            ECNService, "bulk_create_bom_changes",
            new=AsyncMock(side_effect=ECNNotFound(_ECN_ID)),
        ):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/bom-changes/bulk",
                files={"file": ("bom.xlsx", self._one_row_xlsx(),
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        assert resp.status_code == 404

    def test_edit_blocked_at_dc_approved_returns_409(self, client):
        with patch.object(
            ECNService, "bulk_create_bom_changes",
            new=AsyncMock(side_effect=ECNValidationError(
                "BOM changes cannot be edited once the ECN has reached DC_APPROVED (DC role only)"
            )),
        ):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/bom-changes/bulk",
                files={"file": ("bom.xlsx", self._one_row_xlsx(),
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        assert resp.status_code == 409

    def test_parent_not_on_ecn_is_accepted_as_bom_only_change(self, client):
        """ADR-014 — a parent that is NOT an item on this ECN is no longer an
        error. Stargile's ZECNBOMS rows carry their own BMPRNO and never
        referenced the items table; the check requiring the parent to be on
        the ECN was written there and deliberately commented out. The row is
        created standing alone (ecn_item_id NULL).

        This replaces the previous test_unresolved_item_number_returns_422,
        which asserted the constraint this ADR removes.
        """
        created = [_make_change("item-x", "LF200010", item_number="LFAM050001")]
        with patch.object(
            ECNService, "bulk_create_bom_changes", new=AsyncMock(return_value=created)
        ):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/bom-changes/bulk",
                files={"file": ("bom.xlsx", self._one_row_xlsx(),
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        assert resp.status_code == 201
        assert resp.json()[0]["parent_item_number"] == "LFAM050001"

    def test_parent_missing_from_movex_returns_422(self, client):
        """ADR-014 — the rule Stargile actually enforced
        (RequestECNBoMDetailValidationHelper.java:342-352): the parent must
        exist in Movex. Off-ECN parents get an ERP existence check.

        get_item returns {} for a nonexistent item: the MI route reports
        not-found as HTTP 422 / success:false, absorbed by the adapter.
        Verified live against CONO=300 (2026-08-25)."""

        class _MissingParentERP:
            async def get_item(self, item_number: str) -> dict:
                return {}

        app.dependency_overrides[_get_erp_adapter] = lambda: _MissingParentERP()
        resp = client.post(
            f"/api/v1/ecn/{_ECN_ID}/bom-changes/bulk",
            files={"file": ("bom.xlsx", self._one_row_xlsx(),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 422
        assert "does not exist in Movex" in resp.json()["detail"]
