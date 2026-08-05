"""
OSKAR — BOM compare endpoint tests (Slice D, ADR-012 D5).

POST /api/v1/bom/compare                    left/right descriptors + options
GET  /api/v1/bom/comparisons/{id}            fetch a saved comparison
POST /api/v1/bom/compare/upload              multipart upload -> customer-BOM compare
GET  /api/v1/bom/comparisons/{id}/export     xlsx export, fixed field set

Strategy matches tests/routers/test_mpn_search.py / test_parts_alias.py:
TestClient against the real app, service-layer functions patched at their
src.routers.bom import path (not the origin module) — no DB, no HTTP touched
in this test file. get_session is overridden to a stub since every route
here goes through the service layer, which is itself patched.
"""
from __future__ import annotations

import io
from pathlib import Path
from unittest.mock import AsyncMock, patch

import openpyxl
import pytest
from fastapi.testclient import TestClient

from src.adapters.erp.movex import MovexRestAdapter
from src.auth.dependencies import CurrentUser, get_current_user
from src.db import get_session
from src.main import app
from src.services.bom.comparisons import BOMComparison

# Seed app.state with a bare adapter instance so _get_erp_adapter resolves
# without starting the real lifespan — same convention as
# tests/routers/test_bom_browse.py / test_parts_alias.py.
app.state.erp_adapter = MovexRestAdapter.__new__(MovexRestAdapter)

_ENGINEER = CurrentUser(
    username="eng_user",
    display_name="Test Engineer",
    email="eng@scanfil.com",
    groups=["ecn-initiator"],
    jti="test-jti-bom-compare-001",
)


def _make_client(user: CurrentUser) -> TestClient:
    app.dependency_overrides[get_current_user] = lambda: user
    app.dependency_overrides[get_session] = lambda: None
    return TestClient(app, raise_server_exceptions=False)


@pytest.fixture(autouse=True)
def _clear_overrides():
    yield
    app.dependency_overrides.clear()


_SAVED_COMPARISON = BOMComparison(
    id="11111111-1111-1111-1111-111111111111",
    left_descriptor={"type": "erp", "item_number": "LF100001", "facility": "D"},
    right_descriptor={"type": "erp", "item_number": "LF100002", "facility": "D"},
    comparison_result={
        "added": [], "removed": [], "changed": [], "unresolved": [],
        "stats": {"left_count": 0, "right_count": 0, "added_count": 0,
                  "removed_count": 0, "changed_count": 0, "unresolved_count": 0},
    },
    cost_impact=None,
    risk_flags=[],
    created_by="eng_user",
    created_at=__import__("datetime").datetime(2026, 8, 1, tzinfo=__import__("datetime").timezone.utc),
)


class TestGetComparison:
    def test_existing_comparison_returns_200(self):
        with patch("src.routers.bom.get_comparison", new_callable=AsyncMock) as mock:
            mock.return_value = _SAVED_COMPARISON
            client = _make_client(_ENGINEER)
            resp = client.get(f"/api/v1/bom/comparisons/{_SAVED_COMPARISON.id}")

        assert resp.status_code == 200

    def test_response_includes_comparison_result(self):
        with patch("src.routers.bom.get_comparison", new_callable=AsyncMock) as mock:
            mock.return_value = _SAVED_COMPARISON
            client = _make_client(_ENGINEER)
            resp = client.get(f"/api/v1/bom/comparisons/{_SAVED_COMPARISON.id}")

        body = resp.json()
        assert body["id"] == _SAVED_COMPARISON.id
        assert body["comparison_result"]["stats"]["left_count"] == 0

    def test_unknown_id_returns_404(self):
        with patch("src.routers.bom.get_comparison", new_callable=AsyncMock) as mock:
            mock.return_value = None
            client = _make_client(_ENGINEER)
            resp = client.get("/api/v1/bom/comparisons/99999999-9999-9999-9999-999999999999")

        assert resp.status_code == 404

    def test_requires_authentication(self):
        app.dependency_overrides[get_session] = lambda: None
        client = TestClient(app, raise_server_exceptions=False)
        resp = client.get(f"/api/v1/bom/comparisons/{_SAVED_COMPARISON.id}")

        assert resp.status_code == 401


# ---------------------------------------------------------------------------
# POST /api/v1/bom/compare — left/right descriptors + options {key, fields[]}
# ---------------------------------------------------------------------------


class TestPostCompareERPvsERP:
    def test_erp_vs_erp_descriptors_returns_200(self):
        with patch("src.routers.bom.get_single_level_bom", new_callable=AsyncMock) as mock_bom, \
             patch("src.routers.bom.insert_comparison", new_callable=AsyncMock) as mock_insert:
            from src.services.bom.models import BOMHead, BOMLine

            mock_bom.side_effect = [
                BOMHead(item_number="LF100001", structure_type="001", facility="D",
                         description="A", lines=[
                             BOMLine(sequence_number=10, component_number="LF200010",
                                     description="R", operation_number=10, quantity=4.0,
                                     unit_of_measure="EA", from_date=20240101, to_date=99999999),
                         ]),
                BOMHead(item_number="LF100002", structure_type="001", facility="D",
                         description="B", lines=[
                             BOMLine(sequence_number=10, component_number="LF200010",
                                     description="R", operation_number=10, quantity=6.0,
                                     unit_of_measure="EA", from_date=20240101, to_date=99999999),
                         ]),
            ]
            mock_insert.return_value = _SAVED_COMPARISON

            client = _make_client(_ENGINEER)
            resp = client.post(
                "/api/v1/bom/compare",
                json={
                    "left": {"type": "erp", "item_number": "LF100001", "facility": "D"},
                    "right": {"type": "erp", "item_number": "LF100002", "facility": "D"},
                    "options": {},
                },
            )

        assert resp.status_code == 200

    def test_computed_diff_is_passed_to_insert_comparison(self):
        """insert_comparison is mocked (its own persistence is covered by
        tests/integration/test_bom_comparisons.py against real Postgres) —
        this test asserts the router actually computed the real diff and
        handed it to persistence, by inspecting the mock's call args rather
        than the (necessarily mocked) response body."""
        with patch("src.routers.bom.get_single_level_bom", new_callable=AsyncMock) as mock_bom, \
             patch("src.routers.bom.insert_comparison", new_callable=AsyncMock) as mock_insert:
            from src.services.bom.models import BOMHead, BOMLine

            mock_bom.side_effect = [
                BOMHead(item_number="LF100001", structure_type="001", facility="D",
                         description="A", lines=[
                             BOMLine(sequence_number=10, component_number="LF200010",
                                     description="R", operation_number=10, quantity=4.0,
                                     unit_of_measure="EA", from_date=20240101, to_date=99999999),
                         ]),
                BOMHead(item_number="LF100002", structure_type="001", facility="D",
                         description="B", lines=[
                             BOMLine(sequence_number=10, component_number="LF200010",
                                     description="R", operation_number=10, quantity=6.0,
                                     unit_of_measure="EA", from_date=20240101, to_date=99999999),
                         ]),
            ]
            mock_insert.return_value = _SAVED_COMPARISON

            client = _make_client(_ENGINEER)
            client.post(
                "/api/v1/bom/compare",
                json={
                    "left": {"type": "erp", "item_number": "LF100001", "facility": "D"},
                    "right": {"type": "erp", "item_number": "LF100002", "facility": "D"},
                    "options": {},
                },
            )

        _, kwargs = mock_insert.call_args
        assert kwargs["comparison_result"]["stats"]["changed_count"] == 1

    def test_persists_the_comparison(self):
        with patch("src.routers.bom.get_single_level_bom", new_callable=AsyncMock) as mock_bom, \
             patch("src.routers.bom.insert_comparison", new_callable=AsyncMock) as mock_insert:
            from src.services.bom.models import BOMHead

            mock_bom.side_effect = [
                BOMHead(item_number="LF100001", structure_type="001", facility="D", description="A", lines=[]),
                BOMHead(item_number="LF100002", structure_type="001", facility="D", description="B", lines=[]),
            ]
            mock_insert.return_value = _SAVED_COMPARISON

            client = _make_client(_ENGINEER)
            client.post(
                "/api/v1/bom/compare",
                json={
                    "left": {"type": "erp", "item_number": "LF100001", "facility": "D"},
                    "right": {"type": "erp", "item_number": "LF100002", "facility": "D"},
                    "options": {},
                },
            )

        assert mock_insert.await_count == 1

    def test_custom_key_and_fields_options_are_honoured(self):
        with patch("src.routers.bom.get_single_level_bom", new_callable=AsyncMock) as mock_bom, \
             patch("src.routers.bom.insert_comparison", new_callable=AsyncMock) as mock_insert:
            from src.services.bom.models import BOMHead, BOMLine

            mock_bom.side_effect = [
                BOMHead(item_number="LF100001", structure_type="001", facility="D",
                         description="A", lines=[
                             BOMLine(sequence_number=10, component_number="LF200010",
                                     description="Old desc", operation_number=10, quantity=4.0,
                                     unit_of_measure="EA", from_date=20240101, to_date=99999999),
                         ]),
                BOMHead(item_number="LF100002", structure_type="001", facility="D",
                         description="B", lines=[
                             BOMLine(sequence_number=10, component_number="LF200010",
                                     description="New desc", operation_number=20, quantity=4.0,
                                     unit_of_measure="EA", from_date=20240101, to_date=99999999),
                         ]),
            ]
            mock_insert.return_value = _SAVED_COMPARISON

            client = _make_client(_ENGINEER)
            client.post(
                "/api/v1/bom/compare",
                json={
                    "left": {"type": "erp", "item_number": "LF100001", "facility": "D"},
                    "right": {"type": "erp", "item_number": "LF100002", "facility": "D"},
                    "options": {"key": ["component_number"], "fields": ["operation_number"]},
                },
            )

        # component_number-only key matches the op-moved line -> 1 changed
        _, kwargs = mock_insert.call_args
        assert kwargs["comparison_result"]["stats"]["changed_count"] == 1

    def test_unknown_descriptor_type_returns_422(self):
        client = _make_client(_ENGINEER)
        resp = client.post(
            "/api/v1/bom/compare",
            json={
                "left": {"type": "carrier_pigeon", "item_number": "LF100001"},
                "right": {"type": "erp", "item_number": "LF100002", "facility": "D"},
                "options": {},
            },
        )

        assert resp.status_code == 422

    def test_erp_left_not_found_returns_404(self):
        with patch("src.routers.bom.get_single_level_bom", new_callable=AsyncMock) as mock_bom:
            from src.adapters.erp.base import BOMNotFound

            mock_bom.side_effect = BOMNotFound("no BOM for LF999999")

            client = _make_client(_ENGINEER)
            resp = client.post(
                "/api/v1/bom/compare",
                json={
                    "left": {"type": "erp", "item_number": "LF999999", "facility": "D"},
                    "right": {"type": "erp", "item_number": "LF100002", "facility": "D"},
                    "options": {},
                },
            )

        assert resp.status_code == 404


class TestPostCompareSnapshotSide:
    def test_snapshot_descriptor_resolves_via_get_snapshot(self):
        with patch("src.routers.bom.get_snapshot", new_callable=AsyncMock) as mock_snap, \
             patch("src.routers.bom.get_single_level_bom", new_callable=AsyncMock) as mock_bom, \
             patch("src.routers.bom.insert_comparison", new_callable=AsyncMock) as mock_insert:
            from src.services.bom.models import BOMHead, BOMLine
            from src.services.bom.snapshots import BOMSnapshot
            import datetime

            mock_snap.return_value = BOMSnapshot(
                id="snap-1", item_number="LF100001", facility="D", structure_type="001",
                level_mode="single",
                lines=[{"component_number": "LF200010", "operation_number": 10, "quantity": 4.0}],
                line_count=1, content_hash="x" * 64, reason="manual", ecn_id=None,
                captured_by="eng_user", captured_at=datetime.datetime.now(datetime.timezone.utc),
            )
            mock_bom.return_value = BOMHead(
                item_number="LF100002", structure_type="001", facility="D", description="B", lines=[
                    BOMLine(sequence_number=10, component_number="LF200010",
                            description="R", operation_number=10, quantity=4.0,
                            unit_of_measure="EA", from_date=20240101, to_date=99999999),
                ],
            )
            mock_insert.return_value = _SAVED_COMPARISON

            client = _make_client(_ENGINEER)
            resp = client.post(
                "/api/v1/bom/compare",
                json={
                    "left": {"type": "snapshot", "snapshot_id": "snap-1"},
                    "right": {"type": "erp", "item_number": "LF100002", "facility": "D"},
                    "options": {},
                },
            )

        assert resp.status_code == 200
        mock_snap.assert_awaited_once()

    def test_unknown_snapshot_id_returns_404(self):
        with patch("src.routers.bom.get_snapshot", new_callable=AsyncMock) as mock_snap:
            mock_snap.return_value = None

            client = _make_client(_ENGINEER)
            resp = client.post(
                "/api/v1/bom/compare",
                json={
                    "left": {"type": "snapshot", "snapshot_id": "unknown-id"},
                    "right": {"type": "erp", "item_number": "LF100002", "facility": "D"},
                    "options": {},
                },
            )

        assert resp.status_code == 404


# ---------------------------------------------------------------------------
# POST /api/v1/bom/compare/upload — multipart customer-BOM upload vs ERP item
# ---------------------------------------------------------------------------

_FIXTURES_DIR = Path(__file__).parent.parent / "fixtures" / "bom"

_UPLOAD_HEADERS = ["IPN", "CPN", "MFR1", "MPN1", "MFR2", "MPN2", "Designator", "Description", "Quantity", "Footprint"]


def _make_upload_csv(rows: list[dict]) -> bytes:
    lines = [",".join(_UPLOAD_HEADERS)]
    for row in rows:
        lines.append(",".join(str(row.get(h, "")) for h in _UPLOAD_HEADERS))
    return "\n".join(lines).encode("utf-8")


class TestPostCompareUpload:
    def test_valid_upload_returns_200(self):
        csv_bytes = _make_upload_csv([
            {"IPN": "LF200010", "CPN": "CPN-1001", "MFR1": "STMicroelectronics",
             "MPN1": "STM32F103C8T6", "Designator": "U1", "Description": "MCU", "Quantity": "1"},
        ])

        with patch("src.routers.bom.get_single_level_bom", new_callable=AsyncMock) as mock_bom, \
             patch("src.routers.bom.insert_comparison", new_callable=AsyncMock) as mock_insert, \
             patch("src.routers.bom.MovexRestAdapter.lookup_by_alias", new_callable=AsyncMock) as mock_alias:
            from src.services.bom.models import BOMHead, BOMLine

            mock_bom.return_value = BOMHead(
                item_number="LF100001", structure_type="001", facility="D", description="A",
                lines=[
                    BOMLine(sequence_number=10, component_number="LF200010", description="MCU",
                            operation_number=10, quantity=1.0, unit_of_measure="EA",
                            from_date=20240101, to_date=99999999),
                ],
            )
            mock_alias.return_value = [{"ITNO": "LF200010", "POPN": "CPN-1001", "ALWT": "1", "ALWQ": "", "E0PA": ""}]
            mock_insert.return_value = _SAVED_COMPARISON

            client = _make_client(_ENGINEER)
            resp = client.post(
                "/api/v1/bom/compare/upload",
                files={"file": ("customer_bom.csv", csv_bytes, "text/csv")},
                data={"item_number": "LF100001", "facility": "D"},
            )

        assert resp.status_code == 200

    def test_real_customer_bom_csv_fixture_uploads_successfully(self):
        """End-to-end against the actual Slice 0 fixture (multi-row-per-IPN,
        N/A quantity, blank-MFR/MPN row all present in this file)."""
        csv_bytes = (_FIXTURES_DIR / "customer_bom.csv").read_bytes()

        with patch("src.routers.bom.get_single_level_bom", new_callable=AsyncMock) as mock_bom, \
             patch("src.routers.bom.insert_comparison", new_callable=AsyncMock) as mock_insert, \
             patch("src.routers.bom.MovexRestAdapter.lookup_by_alias", new_callable=AsyncMock) as mock_alias:
            from src.services.bom.models import BOMHead

            mock_bom.return_value = BOMHead(
                item_number="LF100001", structure_type="001", facility="D", description="A", lines=[],
            )
            mock_alias.return_value = []
            mock_insert.return_value = _SAVED_COMPARISON

            client = _make_client(_ENGINEER)
            resp = client.post(
                "/api/v1/bom/compare/upload",
                files={"file": ("customer_bom.csv", csv_bytes, "text/csv")},
                data={"item_number": "LF100001", "facility": "D"},
            )

        assert resp.status_code == 200

    def test_real_customer_bom_xlsx_fixture_uploads_successfully(self):
        """The xlsx fixture has a title row above the real header row —
        parse_bulk_upload's header-detection (first non-blank row) handles
        that; this test proves the upload endpoint end-to-end for xlsx too."""
        xlsx_bytes = (_FIXTURES_DIR / "customer_bom.xlsx").read_bytes()

        with patch("src.routers.bom.get_single_level_bom", new_callable=AsyncMock) as mock_bom, \
             patch("src.routers.bom.insert_comparison", new_callable=AsyncMock) as mock_insert, \
             patch("src.routers.bom.MovexRestAdapter.lookup_by_alias", new_callable=AsyncMock) as mock_alias:
            from src.services.bom.models import BOMHead

            mock_bom.return_value = BOMHead(
                item_number="LF100001", structure_type="001", facility="D", description="A", lines=[],
            )
            mock_alias.return_value = []
            mock_insert.return_value = _SAVED_COMPARISON

            client = _make_client(_ENGINEER)
            resp = client.post(
                "/api/v1/bom/compare/upload",
                files={"file": (
                    "customer_bom.xlsx", xlsx_bytes,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )},
                data={"item_number": "LF100001", "facility": "D"},
            )

        assert resp.status_code == 200

    def test_missing_required_columns_returns_422(self):
        bad_csv = b"Foo,Bar\n1,2\n"

        client = _make_client(_ENGINEER)
        resp = client.post(
            "/api/v1/bom/compare/upload",
            files={"file": ("bad.csv", bad_csv, "text/csv")},
            data={"item_number": "LF100001", "facility": "D"},
        )

        assert resp.status_code == 422

    def test_row_missing_quantity_returns_422_naming_the_row_number(self):
        """Per-row validation (plan: "422 with row numbers on bad rows").
        Quantity presence is required; quantity does NOT have to be numeric
        (defect (b) — a non-numeric quantity like "N/A" is valid data, only
        a genuinely BLANK quantity is a row-level error)."""
        csv_bytes = (
            b"IPN,CPN,MFR1,MPN1,MFR2,MPN2,Designator,Description,Quantity,Footprint\n"
            b"LF200010,CPN-1001,STMicroelectronics,STM32F103C8T6,,,U1,MCU,,LQFP48\n"
        )

        client = _make_client(_ENGINEER)
        resp = client.post(
            "/api/v1/bom/compare/upload",
            files={"file": ("bad_row.csv", csv_bytes, "text/csv")},
            data={"item_number": "LF100001", "facility": "D"},
        )

        assert resp.status_code == 422
        assert "Row 1" in resp.json()["detail"]

    def test_non_numeric_quantity_is_not_a_row_validation_error(self):
        """The defect-(b) regression at the endpoint layer: "N/A" quantity
        is accepted (only a genuinely blank quantity is rejected)."""
        csv_bytes = (
            b"IPN,CPN,MFR1,MPN1,MFR2,MPN2,Designator,Description,Quantity,Footprint\n"
            b"LF200010,CPN-1001,STMicroelectronics,STM32F103C8T6,,,U1,MCU,N/A,LQFP48\n"
        )

        with patch("src.routers.bom.get_single_level_bom", new_callable=AsyncMock) as mock_bom, \
             patch("src.routers.bom.insert_comparison", new_callable=AsyncMock) as mock_insert, \
             patch("src.routers.bom.MovexRestAdapter.lookup_by_alias", new_callable=AsyncMock) as mock_alias:
            from src.services.bom.models import BOMHead

            mock_bom.return_value = BOMHead(
                item_number="LF100001", structure_type="001", facility="D", description="A", lines=[],
            )
            mock_alias.return_value = [{"ITNO": "LF200010", "POPN": "CPN-1001", "ALWT": "1", "ALWQ": "", "E0PA": ""}]
            mock_insert.return_value = _SAVED_COMPARISON

            client = _make_client(_ENGINEER)
            resp = client.post(
                "/api/v1/bom/compare/upload",
                files={"file": ("na_qty.csv", csv_bytes, "text/csv")},
                data={"item_number": "LF100001", "facility": "D"},
            )

        assert resp.status_code == 200


# ---------------------------------------------------------------------------
# GET /api/v1/bom/comparisons/{id}/export — xlsx, fixed field set regardless
# of which fields the compare itself was restricted to (parity with PLM:
# "Export to .xlsx only, always using the full fixed field set regardless of
# on-screen column visibility").
# ---------------------------------------------------------------------------

_EXPORT_COMPARISON = BOMComparison(
    id="22222222-2222-2222-2222-222222222222",
    left_descriptor={"type": "erp", "item_number": "LF100001", "facility": "D"},
    right_descriptor={"type": "erp", "item_number": "LF100002", "facility": "D"},
    comparison_result={
        "added": [{"item_number": "LF200099", "quantity": 1.0}],
        "removed": [{"item_number": "LF200098", "quantity": 2.0}],
        "changed": [
            {
                "key": ["LF200010"],
                "left": {"item_number": "LF200010", "quantity": 4.0},
                "right": {"item_number": "LF200010", "quantity": 6.0},
                "field_changes": [{"field": "quantity", "old_value": 4.0, "new_value": 6.0}],
            },
        ],
        "unresolved": [],
        "stats": {"left_count": 2, "right_count": 2, "added_count": 1,
                  "removed_count": 1, "changed_count": 1, "unresolved_count": 0},
    },
    cost_impact=None,
    risk_flags=[],
    created_by="eng_user",
    created_at=__import__("datetime").datetime(2026, 8, 1, tzinfo=__import__("datetime").timezone.utc),
)


class TestExportComparison:
    def test_existing_comparison_export_returns_200(self):
        with patch("src.routers.bom.get_comparison", new_callable=AsyncMock) as mock:
            mock.return_value = _EXPORT_COMPARISON
            client = _make_client(_ENGINEER)
            resp = client.get(f"/api/v1/bom/comparisons/{_EXPORT_COMPARISON.id}/export")

        assert resp.status_code == 200

    def test_export_returns_xlsx_content_type(self):
        with patch("src.routers.bom.get_comparison", new_callable=AsyncMock) as mock:
            mock.return_value = _EXPORT_COMPARISON
            client = _make_client(_ENGINEER)
            resp = client.get(f"/api/v1/bom/comparisons/{_EXPORT_COMPARISON.id}/export")

        assert resp.headers["content-type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    def test_export_is_a_valid_workbook_with_all_three_diff_sections(self):
        with patch("src.routers.bom.get_comparison", new_callable=AsyncMock) as mock:
            mock.return_value = _EXPORT_COMPARISON
            client = _make_client(_ENGINEER)
            resp = client.get(f"/api/v1/bom/comparisons/{_EXPORT_COMPARISON.id}/export")

        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        statuses = {row[0] for row in rows[1:]}  # skip header row
        assert statuses == {"Added", "Removed", "Changed"}

    def test_export_uses_fixed_field_set_regardless_of_compare_fields_restriction(self):
        """PLM parity: export always uses the full fixed field set, even
        though this comparison's own diff was computed with fields
        restricted (comparison_result still carries whatever the diff
        touched — export presents it via a fixed column layout, not a
        layout that varies with what was toggled at compare time)."""
        with patch("src.routers.bom.get_comparison", new_callable=AsyncMock) as mock:
            mock.return_value = _EXPORT_COMPARISON
            client = _make_client(_ENGINEER)
            resp = client.get(f"/api/v1/bom/comparisons/{_EXPORT_COMPARISON.id}/export")

        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        header = [c.value for c in next(ws.iter_rows(max_row=1))]
        assert header == ["Status", "Key", "Field", "Old Value", "New Value"]

    def test_unknown_comparison_id_export_returns_404(self):
        with patch("src.routers.bom.get_comparison", new_callable=AsyncMock) as mock:
            mock.return_value = None
            client = _make_client(_ENGINEER)
            resp = client.get("/api/v1/bom/comparisons/99999999-9999-9999-9999-999999999999/export")

        assert resp.status_code == 404

    def test_wrong_content_type_returns_422(self):
        client = _make_client(_ENGINEER)
        resp = client.post(
            "/api/v1/bom/compare/upload",
            files={"file": ("data.txt", b"not a spreadsheet", "text/plain")},
            data={"item_number": "LF100001", "facility": "D"},
        )

        assert resp.status_code == 422

    def test_empty_file_returns_422(self):
        empty_csv = (",".join(_UPLOAD_HEADERS) + "\n").encode("utf-8")

        client = _make_client(_ENGINEER)
        resp = client.post(
            "/api/v1/bom/compare/upload",
            files={"file": ("empty.csv", empty_csv, "text/csv")},
            data={"item_number": "LF100001", "facility": "D"},
        )

        assert resp.status_code == 422
