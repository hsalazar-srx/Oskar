"""
OSKAR — Bulk MPN upload endpoint tests

POST /api/v1/ecn/{ecn_id}/mpns/bulk   (multipart/form-data, file=<xlsx|csv>)

Template is the real CAD BOM export shape (C P/N, Manufacturer 1/2, Manufacturer
1/2 Part Number) — see BOM-LI_RFSoC_8X8_GNSS_V2I1.csv. One CSV row expands to
1 or 2 MPN rows (primary default + optional alternate).

Strategy: FastAPI TestClient + ECNService methods patched at method level — no DB.
Mirrors tests/routers/test_ecn_items_bulk.py's structure.

Run with: pytest tests/routers/test_ecn_mpns_bulk.py -v
"""
from __future__ import annotations

import io
from datetime import datetime, timezone
from unittest.mock import AsyncMock, patch

import openpyxl
import pytest
from fastapi.testclient import TestClient

from src.auth.dependencies import CurrentUser, get_current_user
from src.db import get_session
from src.main import app
from src.services.ecn import ECNMPNDetail, ECNService
from src.services.ecn.models import ECNNotFound, ECNValidationError

_NOW = datetime(2026, 7, 16, 10, 0, 0, tzinfo=timezone.utc)

_ENGINEER = CurrentUser(
    username="eng_user",
    display_name="Engineer",
    email="eng@scanfil.com",
    groups=["ecn-initiator"],
    jti="test-jti-mpn-bulk-001",
)

_ECN_ID = "ecn-uuid-mpn-bulk-001"

_HEADERS = [
    "C P/N", "Comment", "Description", "Designator", "Manufacturer 1",
    "Manufacturer 1 Part Number", "Manufacturer 2", "Manufacturer 2 Part Number",
]


def _make_mpn(item_id: str, mpn: str, manufacturer: str, is_default: bool) -> ECNMPNDetail:
    return ECNMPNDetail(
        id=f"mpn-{item_id}-{mpn}",
        ecn_item_id=item_id,
        mpn=mpn,
        manufacturer=manufacturer,
        is_default=is_default,
        alias_written=False,
        msl_level=None,
        lifecycle=None,
        eol_date=None,
        lead_time_weeks=None,
        packaging_type=None,
        do_not_buy=False,
        alt_mpn=None,
        notes=None,
        supplier_data_at=None,
        created_at=_NOW,
    )


def _make_xlsx(rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(_HEADERS)
    for row in rows:
        ws.append([row.get(h, "") for h in _HEADERS])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_csv(rows: list[dict]) -> bytes:
    lines = [",".join(_HEADERS)]
    for row in rows:
        lines.append(",".join(str(row.get(h, "")) for h in _HEADERS))
    return "\n".join(lines).encode("utf-8")


@pytest.fixture()
def client():
    app.dependency_overrides[get_current_user] = lambda: _ENGINEER
    app.dependency_overrides[get_session] = lambda: None
    yield TestClient(app)
    app.dependency_overrides.clear()


# ── Happy path ────────────────────────────────────────────────────────────────

class TestBulkMPNHappyPath:
    def test_primary_manufacturer_only_returns_one_mpn(self, client):
        rows = [{"C P/N": "LFDR410018", "Comment": "IC OPAMP", "Manufacturer 1": "Texas Instruments",
                 "Manufacturer 1 Part Number": "LM741CN/NOPB"}]
        xlsx_bytes = _make_xlsx(rows)
        created = [_make_mpn("item-LFDR410018", "LM741CN/NOPB", "Texas Instruments", True)]

        with patch.object(ECNService, "bulk_create_mpns", new=AsyncMock(return_value=created)):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/mpns/bulk",
                files={"file": ("mpns.xlsx", xlsx_bytes,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 201
        data = resp.json()
        assert len(data) == 1
        assert data[0]["mpn"] == "LM741CN/NOPB"
        assert data[0]["is_default"] is True

    def test_primary_and_alternate_manufacturer_returns_two_mpns(self, client):
        """Manufacturer 2 columns populated -> a second, non-default MPN row."""
        rows = [{"C P/N": "LFDR410018", "Manufacturer 1": "Texas Instruments",
                 "Manufacturer 1 Part Number": "LM741CN/NOPB",
                 "Manufacturer 2": "Analog Devices", "Manufacturer 2 Part Number": "LM741ACN"}]
        xlsx_bytes = _make_xlsx(rows)
        created = [
            _make_mpn("item-LFDR410018", "LM741CN/NOPB", "Texas Instruments", True),
            _make_mpn("item-LFDR410018", "LM741ACN", "Analog Devices", False),
        ]

        with patch.object(ECNService, "bulk_create_mpns", new=AsyncMock(return_value=created)) as mock_bulk:
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/mpns/bulk",
                files={"file": ("mpns.xlsx", xlsx_bytes,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 201
        data = resp.json()
        assert len(data) == 2
        assert data[0]["is_default"] is True
        assert data[1]["is_default"] is False

        # The router must have expanded the one CAD-BOM row into two validated rows.
        forwarded_rows = mock_bulk.call_args.args[1]
        assert len(forwarded_rows) == 2
        assert forwarded_rows[0]["mpn"] == "LM741CN/NOPB"
        assert forwarded_rows[1]["mpn"] == "LM741ACN"

    def test_csv_upload_returns_201(self, client):
        rows = [{"C P/N": "LFDR410018", "Manufacturer 1": "Texas Instruments",
                 "Manufacturer 1 Part Number": "LM741CN/NOPB"}]
        csv_bytes = _make_csv(rows)
        created = [_make_mpn("item-LFDR410018", "LM741CN/NOPB", "Texas Instruments", True)]

        with patch.object(ECNService, "bulk_create_mpns", new=AsyncMock(return_value=created)):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/mpns/bulk",
                files={"file": ("mpns.csv", csv_bytes, "text/csv")},
            )

        assert resp.status_code == 201
        assert len(resp.json()) == 1


# ── Validation errors ─────────────────────────────────────────────────────────

class TestBulkMPNValidationErrors:
    def test_wrong_content_type_returns_422(self, client):
        resp = client.post(
            f"/api/v1/ecn/{_ECN_ID}/mpns/bulk",
            files={"file": ("mpns.txt", b"some text", "text/plain")},
        )
        assert resp.status_code == 422

    def test_file_too_large_returns_413(self, client):
        big_bytes = b"x" * (2 * 1024 * 1024 + 1)
        resp = client.post(
            f"/api/v1/ecn/{_ECN_ID}/mpns/bulk",
            files={"file": ("big.xlsx", big_bytes,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 413

    def test_missing_required_columns_returns_422(self, client):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Wrong Col A", "Wrong Col B"])
        ws.append(["val1", "val2"])
        buf = io.BytesIO()
        wb.save(buf)

        with patch.object(ECNService, "bulk_create_mpns", new=AsyncMock()):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/mpns/bulk",
                files={"file": ("bad.xlsx", buf.getvalue(),
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 422
        assert "missing" in resp.json()["detail"].lower() or "columns" in resp.json()["detail"].lower()

    def test_empty_file_returns_422(self, client):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(_HEADERS)
        buf = io.BytesIO()
        wb.save(buf)

        with patch.object(ECNService, "bulk_create_mpns", new=AsyncMock()):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/mpns/bulk",
                files={"file": ("empty.xlsx", buf.getvalue(),
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 422

    def test_blank_c_pn_rejects_batch_with_422(self, client):
        """MVP: a component row with a real MPN but no C P/N is a hard failure,
        not a silent skip — auto-resolving it against Movex/DigiKey is deferred."""
        rows = [{"C P/N": "", "Manufacturer 1": "Texas Instruments",
                 "Manufacturer 1 Part Number": "LM741CN/NOPB"}]
        xlsx_bytes = _make_xlsx(rows)

        with patch.object(ECNService, "bulk_create_mpns", new=AsyncMock()):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/mpns/bulk",
                files={"file": ("mpns.xlsx", xlsx_bytes,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 422

    def test_fully_blank_filler_row_is_skipped_not_rejected(self, client):
        """A row with no Manufacturer 1 Part Number at all (pure spacer row in
        the CAD export) is silently skipped, unlike a real-but-unresolved row."""
        rows = [
            {"C P/N": "", "Manufacturer 1": "", "Manufacturer 1 Part Number": ""},
            {"C P/N": "LFDR410018", "Manufacturer 1": "Texas Instruments",
             "Manufacturer 1 Part Number": "LM741CN/NOPB"},
        ]
        xlsx_bytes = _make_xlsx(rows)
        created = [_make_mpn("item-LFDR410018", "LM741CN/NOPB", "Texas Instruments", True)]

        with patch.object(ECNService, "bulk_create_mpns", new=AsyncMock(return_value=created)) as mock_bulk:
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/mpns/bulk",
                files={"file": ("mpns.xlsx", xlsx_bytes,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 201
        forwarded_rows = mock_bulk.call_args.args[1]
        assert len(forwarded_rows) == 1

    def test_duplicate_item_mpn_pair_returns_409(self, client):
        rows = [
            {"C P/N": "LFDR410018", "Manufacturer 1": "Texas Instruments",
             "Manufacturer 1 Part Number": "LM741CN/NOPB"},
            {"C P/N": "LFDR410018", "Manufacturer 1": "Texas Instruments (Alt Source)",
             "Manufacturer 1 Part Number": "LM741CN/NOPB"},
        ]
        xlsx_bytes = _make_xlsx(rows)

        with patch.object(ECNService, "bulk_create_mpns", new=AsyncMock()):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/mpns/bulk",
                files={"file": ("mpns.xlsx", xlsx_bytes,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 409


# ── ECN / item state guards ──────────────────────────────────────────────────

class TestBulkMPNStateGuards:
    def _one_row_xlsx(self) -> bytes:
        return _make_xlsx([{"C P/N": "LFDR410018", "Manufacturer 1": "Texas Instruments",
                             "Manufacturer 1 Part Number": "LM741CN/NOPB"}])

    def test_ecn_not_found_returns_404(self, client):
        with patch.object(
            ECNService, "bulk_create_mpns",
            new=AsyncMock(side_effect=ECNNotFound(_ECN_ID)),
        ):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/mpns/bulk",
                files={"file": ("mpns.xlsx", self._one_row_xlsx(),
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        assert resp.status_code == 404

    def test_ecn_not_draft_returns_409(self, client):
        with patch.object(
            ECNService, "bulk_create_mpns",
            new=AsyncMock(side_effect=ECNValidationError("ECN is not in DRAFT status")),
        ):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/mpns/bulk",
                files={"file": ("mpns.xlsx", self._one_row_xlsx(),
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        assert resp.status_code == 409

    def test_unresolved_item_number_returns_422(self, client):
        with patch.object(
            ECNService, "bulk_create_mpns",
            new=AsyncMock(side_effect=ECNValidationError(
                "Row 1: item_number 'LFDR410018' was not found on this ECN — add it via item upload first"
            )),
        ):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/mpns/bulk",
                files={"file": ("mpns.xlsx", self._one_row_xlsx(),
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        assert resp.status_code == 422
        assert "item_number" in resp.json()["detail"]
