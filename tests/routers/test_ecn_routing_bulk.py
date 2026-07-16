"""
OSKAR — Bulk routing operation upload endpoint tests

POST /api/v1/ecn/{ecn_id}/routing/bulk   (multipart/form-data, file=<xlsx|csv>)

Strategy: FastAPI TestClient + ECNService methods patched at method level — no DB.
get_current_user overridden via dependency_overrides. Mirrors
tests/routers/test_ecn_items_bulk.py's structure.

Run with: pytest tests/routers/test_ecn_routing_bulk.py -v
"""
from __future__ import annotations

import io
from datetime import datetime, timezone
from unittest.mock import AsyncMock, patch

import openpyxl
import pytest
from fastapi.testclient import TestClient

from src.auth.dependencies import CurrentUser, get_current_user
from src.db import get_session
from src.main import app
from src.services.ecn import ECNService, RoutingOperationResponse
from src.services.ecn.models import ECNNotFound, ECNValidationError

_NOW = datetime(2026, 7, 16, 10, 0, 0, tzinfo=timezone.utc)

_ENGINEER = CurrentUser(
    username="eng_user",
    display_name="Engineer",
    email="eng@scanfil.com",
    groups=["ecn-initiator"],
    jti="test-jti-routing-bulk-001",
)

_ECN_ID = "ecn-uuid-routing-bulk-001"

_HEADERS = [
    "Item No", "Operation No", "Operation Description", "Work Centre",
    "Run Time", "Setup Time", "Change Type",
]


def _make_op(item_id: str, opno: int, work_centre: str = "FCT", run_time: float = 3.25) -> RoutingOperationResponse:
    return RoutingOperationResponse(
        id=f"op-{item_id}-{opno}",
        ecn_item_id=item_id,
        operation_number=opno,
        operation_description="Functional Test",
        work_centre=work_centre,
        run_time=run_time,
        setup_time=0.1,
        change_type="ADD",
        movex_snapshot=None,
        created_at=_NOW,
        updated_at=_NOW,
    )


def _make_xlsx(rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(_HEADERS)
    for row in rows:
        ws.append([row.get(h, "") for h in _HEADERS])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_csv(rows: list[dict]) -> bytes:
    lines = [",".join(_HEADERS)]
    for row in rows:
        lines.append(",".join(str(row.get(h, "")) for h in _HEADERS))
    return "\n".join(lines).encode("utf-8")


@pytest.fixture()
def client():
    app.dependency_overrides[get_current_user] = lambda: _ENGINEER
    app.dependency_overrides[get_session] = lambda: None
    yield TestClient(app)
    app.dependency_overrides.clear()


# ── Happy path ────────────────────────────────────────────────────────────────

class TestBulkRoutingHappyPath:
    def test_multi_item_upload_returns_201(self, client):
        """Rows spread across two different items in one file."""
        rows = [
            {"Item No": "LFAM050001", "Operation No": "10", "Operation Description": "Kitting",
             "Work Centre": "KIT", "Run Time": "0.83", "Change Type": "ADD"},
            {"Item No": "LFAM050002", "Operation No": "130", "Operation Description": "Functional Test",
             "Work Centre": "FCT", "Run Time": "3.25", "Change Type": "ADD"},
        ]
        xlsx_bytes = _make_xlsx(rows)
        created = [_make_op("item-LFAM050001", 10, "KIT", 0.83), _make_op("item-LFAM050002", 130)]

        with patch.object(ECNService, "bulk_create_routing_operations", new=AsyncMock(return_value=created)):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/routing/bulk",
                files={"file": ("routing.xlsx", xlsx_bytes,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 201
        data = resp.json()
        assert len(data) == 2
        assert data[0]["ecn_item_id"] == "item-LFAM050001"
        assert data[1]["ecn_item_id"] == "item-LFAM050002"

    def test_single_item_full_routing_upload_returns_201(self, client):
        """Many operation rows for the SAME item — the Labour Routing.csv scenario."""
        rows = [
            {"Item No": "LFAM050001", "Operation No": "10", "Operation Description": "Kitting",
             "Work Centre": "KIT", "Run Time": "0.83", "Change Type": "ADD"},
            {"Item No": "LFAM050001", "Operation No": "20", "Operation Description": "Machine Labelling",
             "Work Centre": "LABEL", "Run Time": "0.15", "Change Type": "ADD"},
            {"Item No": "LFAM050001", "Operation No": "130", "Operation Description": "Functional Test",
             "Work Centre": "FCT", "Run Time": "3.25", "Change Type": "ADD"},
        ]
        xlsx_bytes = _make_xlsx(rows)
        created = [
            _make_op("item-LFAM050001", 10, "KIT", 0.83),
            _make_op("item-LFAM050001", 20, "LABEL", 0.15),
            _make_op("item-LFAM050001", 130),
        ]

        with patch.object(ECNService, "bulk_create_routing_operations", new=AsyncMock(return_value=created)):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/routing/bulk",
                files={"file": ("routing.xlsx", xlsx_bytes,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 201
        data = resp.json()
        assert len(data) == 3
        assert all(op["ecn_item_id"] == "item-LFAM050001" for op in data)

    def test_csv_upload_returns_201(self, client):
        rows = [{"Item No": "LFAM050001", "Operation No": "130", "Operation Description": "Functional Test",
                 "Work Centre": "FCT", "Run Time": "3.25", "Change Type": "ADD"}]
        csv_bytes = _make_csv(rows)
        created = [_make_op("item-LFAM050001", 130)]

        with patch.object(ECNService, "bulk_create_routing_operations", new=AsyncMock(return_value=created)):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/routing/bulk",
                files={"file": ("routing.csv", csv_bytes, "text/csv")},
            )

        assert resp.status_code == 201
        assert len(resp.json()) == 1


# ── Validation errors ─────────────────────────────────────────────────────────

class TestBulkRoutingValidationErrors:
    def test_wrong_content_type_returns_422(self, client):
        resp = client.post(
            f"/api/v1/ecn/{_ECN_ID}/routing/bulk",
            files={"file": ("routing.txt", b"some text", "text/plain")},
        )
        assert resp.status_code == 422

    def test_file_too_large_returns_413(self, client):
        big_bytes = b"x" * (2 * 1024 * 1024 + 1)
        resp = client.post(
            f"/api/v1/ecn/{_ECN_ID}/routing/bulk",
            files={"file": ("big.xlsx", big_bytes,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 413

    def test_missing_required_columns_returns_422(self, client):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Wrong Col A", "Wrong Col B"])
        ws.append(["val1", "val2"])
        buf = io.BytesIO()
        wb.save(buf)

        with patch.object(ECNService, "bulk_create_routing_operations", new=AsyncMock()):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/routing/bulk",
                files={"file": ("bad.xlsx", buf.getvalue(),
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 422
        assert "missing" in resp.json()["detail"].lower() or "columns" in resp.json()["detail"].lower()

    def test_empty_file_returns_422(self, client):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(_HEADERS)
        buf = io.BytesIO()
        wb.save(buf)

        with patch.object(ECNService, "bulk_create_routing_operations", new=AsyncMock()):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/routing/bulk",
                files={"file": ("empty.xlsx", buf.getvalue(),
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 422

    def test_invalid_change_type_returns_422(self, client):
        rows = [{"Item No": "LFAM050001", "Operation No": "130", "Operation Description": "Functional Test",
                 "Work Centre": "FCT", "Run Time": "3.25", "Change Type": "BOGUS"}]
        xlsx_bytes = _make_xlsx(rows)

        with patch.object(ECNService, "bulk_create_routing_operations", new=AsyncMock()):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/routing/bulk",
                files={"file": ("routing.xlsx", xlsx_bytes,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 422

    def test_duplicate_item_operation_pair_returns_409(self, client):
        rows = [
            {"Item No": "LFAM050001", "Operation No": "130", "Operation Description": "Functional Test",
             "Work Centre": "FCT", "Run Time": "3.25", "Change Type": "ADD"},
            {"Item No": "LFAM050001", "Operation No": "130", "Operation Description": "Duplicate",
             "Work Centre": "FCT", "Run Time": "3.25", "Change Type": "ADD"},
        ]
        xlsx_bytes = _make_xlsx(rows)

        with patch.object(ECNService, "bulk_create_routing_operations", new=AsyncMock()):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/routing/bulk",
                files={"file": ("routing.xlsx", xlsx_bytes,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 409


# ── ECN / item state guards ──────────────────────────────────────────────────

class TestBulkRoutingStateGuards:
    def _one_row_xlsx(self) -> bytes:
        return _make_xlsx([{"Item No": "LFAM050001", "Operation No": "130",
                             "Operation Description": "Functional Test", "Work Centre": "FCT",
                             "Run Time": "3.25", "Change Type": "ADD"}])

    def test_ecn_not_found_returns_404(self, client):
        with patch.object(
            ECNService, "bulk_create_routing_operations",
            new=AsyncMock(side_effect=ECNNotFound(_ECN_ID)),
        ):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/routing/bulk",
                files={"file": ("routing.xlsx", self._one_row_xlsx(),
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        assert resp.status_code == 404

    def test_ecn_not_draft_returns_409(self, client):
        with patch.object(
            ECNService, "bulk_create_routing_operations",
            new=AsyncMock(side_effect=ECNValidationError("ECN is not in DRAFT status")),
        ):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/routing/bulk",
                files={"file": ("routing.xlsx", self._one_row_xlsx(),
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        assert resp.status_code == 409

    def test_unresolved_item_number_returns_422(self, client):
        """item_number not already on this ECN — bulk routing does not create items."""
        with patch.object(
            ECNService, "bulk_create_routing_operations",
            new=AsyncMock(side_effect=ECNValidationError(
                "Row 1: item_number 'LFAM050001' was not found on this ECN — add it via item upload first"
            )),
        ):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/routing/bulk",
                files={"file": ("routing.xlsx", self._one_row_xlsx(),
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        assert resp.status_code == 422
        assert "item_number" in resp.json()["detail"]
