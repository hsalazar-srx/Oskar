"""
OSKAR — bulk_export.py unit tests: BulkExportSpec / ExportColumn / build_xlsx.

Pure function, no DB, no FastAPI — verifies the xlsx bytes round-trip
correctly via openpyxl for both dict-shaped and object-shaped rows.
"""
from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO

import openpyxl
import pytest

from src.routers.bulk_export import BulkExportSpec, ExportColumn, build_xlsx


def _load(xlsx_bytes: bytes):
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes))
    ws = wb.active
    return ws, wb


class TestBuildXlsxWithDictRows:
    def test_header_row_matches_column_headers(self):
        spec = BulkExportSpec(
            sheet_name="Items",
            columns=[
                ExportColumn(header="Item No", getter=lambda r: r["item_number"]),
                ExportColumn(header="Item Name", getter=lambda r: r["item_name"]),
            ],
        )
        xlsx = build_xlsx([{"item_number": "LF-001", "item_name": "Widget"}], spec)
        ws, wb = _load(xlsx)
        assert [c.value for c in ws[1]] == ["Item No", "Item Name"]
        wb.close()

    def test_data_rows_written_in_order(self):
        spec = BulkExportSpec(
            sheet_name="Items",
            columns=[ExportColumn(header="Item No", getter=lambda r: r["item_number"])],
        )
        rows = [{"item_number": "LF-001"}, {"item_number": "LF-002"}]
        xlsx = build_xlsx(rows, spec)
        ws, wb = _load(xlsx)
        assert ws.cell(row=2, column=1).value == "LF-001"
        assert ws.cell(row=3, column=1).value == "LF-002"
        wb.close()

    def test_empty_rows_produces_header_only(self):
        spec = BulkExportSpec(
            sheet_name="Items",
            columns=[ExportColumn(header="Item No", getter=lambda r: r["item_number"])],
        )
        xlsx = build_xlsx([], spec)
        ws, wb = _load(xlsx)
        assert ws.max_row == 1
        assert ws.cell(row=1, column=1).value == "Item No"
        wb.close()

    def test_sheet_name_applied(self):
        spec = BulkExportSpec(
            sheet_name="Routing Ops",
            columns=[ExportColumn(header="Op No", getter=lambda r: r["op_no"])],
        )
        xlsx = build_xlsx([], spec)
        ws, wb = _load(xlsx)
        assert ws.title == "Routing Ops"
        wb.close()


class TestBuildXlsxWithDataclassRows:
    def test_getter_can_read_dataclass_attributes(self):
        @dataclass
        class Row:
            mpn: str
            manufacturer: str | None

        spec = BulkExportSpec(
            sheet_name="MPNs",
            columns=[
                ExportColumn(header="MPN", getter=lambda r: r.mpn),
                ExportColumn(header="Manufacturer", getter=lambda r: r.manufacturer),
            ],
        )
        xlsx = build_xlsx([Row(mpn="ABC123", manufacturer="Yageo")], spec)
        ws, wb = _load(xlsx)
        assert ws.cell(row=2, column=1).value == "ABC123"
        assert ws.cell(row=2, column=2).value == "Yageo"
        wb.close()


class TestBuildXlsxReturnType:
    def test_returns_bytes(self):
        spec = BulkExportSpec(
            sheet_name="Items",
            columns=[ExportColumn(header="Item No", getter=lambda r: r["item_number"])],
        )
        result = build_xlsx([], spec)
        assert isinstance(result, bytes)
        assert len(result) > 0
