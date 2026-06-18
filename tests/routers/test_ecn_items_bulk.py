"""
OSKAR — Bulk item upload endpoint tests

POST /api/v1/ecn/{ecn_id}/items/bulk   (multipart/form-data, file=<xlsx|csv>)

Strategy: FastAPI TestClient + ECNService methods patched at method level — no DB.
get_current_user overridden via dependency_overrides.

TDD: written before service/router implementation.

Run with: pytest tests/routers/test_ecn_items_bulk.py -v
"""
from __future__ import annotations

import io
from datetime import datetime, timezone
from unittest.mock import AsyncMock, patch

import openpyxl
import pytest
from fastapi.testclient import TestClient

from src.auth.dependencies import CurrentUser, get_current_user
from src.db import get_session
from src.main import app
from src.services.ecn import ECNItemDetail, ECNMPNDetail, ECNService
from src.services.ecn.models import ECNNotFound, ECNValidationError

_NOW = datetime(2026, 6, 17, 10, 0, 0, tzinfo=timezone.utc)

_ENGINEER = CurrentUser(
    username="eng_user",
    display_name="Engineer",
    email="eng@scanfil.com",
    groups=["OSKAR-Engineers"],
    jti="test-jti-bulk-001",
)

_ECN_ID = "ecn-uuid-bulk-001"


def _make_item(item_number: str, line_number: int) -> ECNItemDetail:
    return ECNItemDetail(
        id=f"item-{item_number}",
        ecn_id=_ECN_ID,
        line_number=line_number,
        is_new_item=True,
        item_number=item_number,
        item_name=f"Item {item_number}",
        description_2=None,
        drawing_number=None,
        drawing_created=False,
        procurement_group="PAS",
        product_group="RES",
        unit_of_measure="EA",
        item_group=None,
        customer_alias=None,
        customer_part_number=None,
        effectivity_type="IMMEDIATE",
        effectivity_from=None,
        created_at=_NOW,
        updated_at=_NOW,
        mpns=[],
    )


def _make_xlsx(rows: list[dict]) -> bytes:
    """Build a minimal xlsx matching the Oskar bulk upload template schema."""
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        "ECN Line", "Is New Item", "Item No", "Item Name", "Item Status",
        "Item Description", "Drawing No", "Procurement Group", "Product Group",
        "Item Group", "Unit Of Measurement", "Revision No", "Supplier",
        "Responsible", "Customer Alias", "Order Type", "Lead Free Code",
        "Good Receiving Method",
    ]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_csv(rows: list[dict]) -> bytes:
    headers = [
        "ECN Line", "Is New Item", "Item No", "Item Name", "Item Status",
        "Item Description", "Drawing No", "Procurement Group", "Product Group",
        "Item Group", "Unit Of Measurement", "Revision No", "Supplier",
        "Responsible", "Customer Alias", "Order Type", "Lead Free Code",
        "Good Receiving Method",
    ]
    lines = [",".join(headers)]
    for row in rows:
        lines.append(",".join(str(row.get(h, "")) for h in headers))
    return "\n".join(lines).encode("utf-8")


@pytest.fixture()
def client():
    app.dependency_overrides[get_current_user] = lambda: _ENGINEER
    app.dependency_overrides[get_session] = lambda: None
    yield TestClient(app)
    app.dependency_overrides.clear()


# ── Happy path ────────────────────────────────────────────────────────────────

class TestBulkCreateHappyPath:
    def test_xlsx_upload_returns_201_with_items(self, client):
        rows = [
            {"Is New Item": "1", "Item No": "LFSC691234", "Item Name": "RESISTOR SMD",
             "Item Status": "20", "Procurement Group": "PAS", "Product Group": "RES",
             "Unit Of Measurement": "EA", "Order Type": "010", "Lead Free Code": "PBF",
             "Good Receiving Method": "010"},
            {"Is New Item": "1", "Item No": "LFSC695678", "Item Name": "CAPACITOR SMD",
             "Item Status": "20", "Procurement Group": "PAS", "Product Group": "CAPS",
             "Unit Of Measurement": "EA", "Order Type": "010", "Lead Free Code": "PBF",
             "Good Receiving Method": "010"},
        ]
        xlsx_bytes = _make_xlsx(rows)
        created = [_make_item("LFSC691234", 1), _make_item("LFSC695678", 2)]

        with patch.object(ECNService, "bulk_create_items", new=AsyncMock(return_value=created)):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/items/bulk",
                files={"file": ("items.xlsx", xlsx_bytes,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 201
        data = resp.json()
        assert isinstance(data, list)
        assert len(data) == 2
        assert data[0]["item_number"] == "LFSC691234"
        assert data[1]["item_number"] == "LFSC695678"

    def test_csv_upload_returns_201(self, client):
        rows = [
            {"Is New Item": "1", "Item No": "LFSC691234", "Item Name": "RESISTOR SMD",
             "Item Status": "20", "Procurement Group": "PAS", "Product Group": "RES",
             "Unit Of Measurement": "EA", "Order Type": "010", "Lead Free Code": "PBF",
             "Good Receiving Method": "010"},
        ]
        csv_bytes = _make_csv(rows)
        created = [_make_item("LFSC691234", 1)]

        with patch.object(ECNService, "bulk_create_items", new=AsyncMock(return_value=created)):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/items/bulk",
                files={"file": ("items.csv", csv_bytes, "text/csv")},
            )

        assert resp.status_code == 201
        assert len(resp.json()) == 1

    def test_response_contains_all_ecn_item_fields(self, client):
        rows = [{"Is New Item": "1", "Item No": "LFSC691234", "Item Name": "RESISTOR SMD",
                 "Item Status": "20", "Procurement Group": "PAS", "Product Group": "RES",
                 "Unit Of Measurement": "EA", "Order Type": "010", "Lead Free Code": "PBF",
                 "Good Receiving Method": "010"}]
        xlsx_bytes = _make_xlsx(rows)
        created = [_make_item("LFSC691234", 1)]

        with patch.object(ECNService, "bulk_create_items", new=AsyncMock(return_value=created)):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/items/bulk",
                files={"file": ("items.xlsx", xlsx_bytes,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        item = resp.json()[0]
        assert "id" in item
        assert "ecn_id" in item
        assert "line_number" in item
        assert "item_number" in item
        assert "effectivity_type" in item
        assert "mpns" in item


# ── Validation errors ─────────────────────────────────────────────────────────

class TestBulkCreateValidationErrors:
    def test_wrong_content_type_returns_422(self, client):
        resp = client.post(
            f"/api/v1/ecn/{_ECN_ID}/items/bulk",
            files={"file": ("items.txt", b"some text", "text/plain")},
        )
        assert resp.status_code == 422
        detail = resp.json()["detail"]
        assert any("content type" in str(e).lower() or "xlsx" in str(e).lower() for e in (
            [detail] if isinstance(detail, str) else detail
        ))

    def test_file_too_large_returns_413(self, client):
        big_bytes = b"x" * (2 * 1024 * 1024 + 1)  # 2 MB+
        resp = client.post(
            f"/api/v1/ecn/{_ECN_ID}/items/bulk",
            files={"file": ("big.xlsx", big_bytes,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 413

    def test_missing_required_columns_returns_422(self, client):
        # File with wrong column names (no Item No, no Item Name)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Wrong Col A", "Wrong Col B"])
        ws.append(["val1", "val2"])
        buf = io.BytesIO()
        wb.save(buf)

        with patch.object(ECNService, "bulk_create_items", new=AsyncMock()):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/items/bulk",
                files={"file": ("bad.xlsx", buf.getvalue(),
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 422
        assert "missing" in resp.json()["detail"].lower() or "columns" in resp.json()["detail"].lower()

    def test_empty_file_returns_422(self, client):
        wb = openpyxl.Workbook()
        ws = wb.active
        # Header row only — no data rows
        ws.append(["ECN Line", "Is New Item", "Item No", "Item Name", "Item Description",
                   "Drawing No", "Procurement Group", "Product Group", "Item Group",
                   "Unit Of Measurement", "Revision No", "Supplier", "Responsible", "Customer Alias"])
        buf = io.BytesIO()
        wb.save(buf)

        with patch.object(ECNService, "bulk_create_items", new=AsyncMock()):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/items/bulk",
                files={"file": ("empty.xlsx", buf.getvalue(),
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 422

    def test_service_validation_error_returns_422_with_row_details(self, client):
        rows = [{"Is New Item": "1", "Item No": "LFSC691234", "Item Name": "X" * 31,
                 "Item Status": "20", "Procurement Group": "PAS", "Product Group": "RES",
                 "Unit Of Measurement": "EA", "Order Type": "010", "Lead Free Code": "PBF",
                 "Good Receiving Method": "010"}]
        xlsx_bytes = _make_xlsx(rows)

        with patch.object(
            ECNService, "bulk_create_items",
            new=AsyncMock(side_effect=ECNValidationError("Row 1: item_name exceeds 30 characters")),
        ):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/items/bulk",
                files={"file": ("items.xlsx", xlsx_bytes,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 422
        assert "row" in resp.json()["detail"].lower() or "item_name" in resp.json()["detail"].lower()


# ── ECN state guards ──────────────────────────────────────────────────────────

class TestBulkCreateStateGuards:
    def test_ecn_not_found_returns_404(self, client):
        rows = [{"Is New Item": "1", "Item No": "LFSC691234", "Item Name": "RESISTOR SMD",
                 "Item Status": "20", "Procurement Group": "PAS", "Product Group": "RES",
                 "Unit Of Measurement": "EA", "Order Type": "010", "Lead Free Code": "PBF",
                 "Good Receiving Method": "010"}]
        xlsx_bytes = _make_xlsx(rows)

        with patch.object(
            ECNService, "bulk_create_items",
            new=AsyncMock(side_effect=ECNNotFound(_ECN_ID)),
        ):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/items/bulk",
                files={"file": ("items.xlsx", xlsx_bytes,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 404

    def test_ecn_not_draft_returns_409(self, client):
        rows = [{"Is New Item": "1", "Item No": "LFSC691234", "Item Name": "RESISTOR SMD",
                 "Item Status": "20", "Procurement Group": "PAS", "Product Group": "RES",
                 "Unit Of Measurement": "EA", "Order Type": "010", "Lead Free Code": "PBF",
                 "Good Receiving Method": "010"}]
        xlsx_bytes = _make_xlsx(rows)

        with patch.object(
            ECNService, "bulk_create_items",
            new=AsyncMock(side_effect=ECNValidationError("ECN is not in DRAFT status")),
        ):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/items/bulk",
                files={"file": ("items.xlsx", xlsx_bytes,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 409

    def test_duplicate_item_number_returns_409(self, client):
        rows = [
            {"Is New Item": "1", "Item No": "LFSC691234", "Item Name": "RESISTOR SMD",
             "Item Status": "20", "Procurement Group": "PAS", "Product Group": "RES",
             "Unit Of Measurement": "EA", "Order Type": "010", "Lead Free Code": "PBF",
             "Good Receiving Method": "010"},
            {"Is New Item": "1", "Item No": "LFSC691234", "Item Name": "DUPLICATE",
             "Item Status": "20", "Procurement Group": "PAS", "Product Group": "RES",
             "Unit Of Measurement": "EA", "Order Type": "010", "Lead Free Code": "PBF",
             "Good Receiving Method": "010"},
        ]
        xlsx_bytes = _make_xlsx(rows)

        # Duplicate within the batch itself — caught by endpoint before calling service
        resp = client.post(
            f"/api/v1/ecn/{_ECN_ID}/items/bulk",
            files={"file": ("items.xlsx", xlsx_bytes,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

        assert resp.status_code == 409


# ── Leading-zero preservation ─────────────────────────────────────────────────

class TestLeadingZeroPreservation:
    def test_item_number_with_leading_zero_preserved(self, client):
        """Regression: SheetJS / openpyxl must not coerce '0034567' to 34567."""
        rows = [{"Is New Item": "1", "Item No": "0034567", "Item Name": "FUSE",
                 "Item Status": "20", "Procurement Group": "MAG", "Product Group": "FUSE",
                 "Unit Of Measurement": "EA", "Order Type": "010", "Lead Free Code": "PBF",
                 "Good Receiving Method": "010"}]
        xlsx_bytes = _make_xlsx(rows)
        created = [_make_item("0034567", 1)]

        captured_items: list = []

        async def capture_bulk(ecn_id, items, *, session=None):
            captured_items.extend(items)
            return created

        with patch.object(ECNService, "bulk_create_items", new=AsyncMock(side_effect=capture_bulk)):
            resp = client.post(
                f"/api/v1/ecn/{_ECN_ID}/items/bulk",
                files={"file": ("items.xlsx", xlsx_bytes,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 201
        # Verify the item_number passed to the service has the leading zero intact
        assert any(i.get("item_number") == "0034567" for i in captured_items)


# ── Authentication ────────────────────────────────────────────────────────────

class TestBulkCreateAuthentication:
    def test_unauthenticated_returns_401(self):
        # No dependency override — uses real get_current_user which requires a token
        app.dependency_overrides[get_session] = lambda: None
        c = TestClient(app, raise_server_exceptions=False)
        rows = [{"Is New Item": "1", "Item No": "LFSC691234", "Item Name": "RESISTOR SMD",
                 "Item Status": "20", "Procurement Group": "PAS", "Product Group": "RES",
                 "Unit Of Measurement": "EA", "Order Type": "010", "Lead Free Code": "PBF",
                 "Good Receiving Method": "010"}]
        xlsx_bytes = _make_xlsx(rows)
        resp = c.post(
            f"/api/v1/ecn/{_ECN_ID}/items/bulk",
            files={"file": ("items.xlsx", xlsx_bytes,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        app.dependency_overrides.clear()
        assert resp.status_code == 401
